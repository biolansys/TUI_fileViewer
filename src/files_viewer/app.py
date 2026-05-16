from __future__ import annotations

import importlib.resources as importlib_resources
import json
import os
import pickle
import pprint
import re
import sqlite3
import webbrowser
import copy
import gzip
import xml.dom.minidom as minidom
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import numpy as np
import pandas as pd
import pyarrow as pa
import pyarrow.dataset as ds
import pyarrow.orc as pa_orc
from bs4 import BeautifulSoup
from fastavro import reader as avro_reader
from fastavro import writer as avro_writer
from PIL import Image, ImageOps
from rich.markdown import Markdown
from rich.text import Text
from pypdf import PdfReader
from textual import on, work
from textual.app import App, ComposeResult
from textual.binding import Binding
from textual.containers import Horizontal, Vertical, VerticalScroll
from textual.message import Message
from textual.screen import ModalScreen
from textual.widgets import Button, DataTable, DirectoryTree, Footer, Header, Input, Label, Static, TabbedContent, TabPane, TextArea, Tree
from textual.widgets.tree import TreeNode

TABULAR_EXTENSIONS = {".csv", ".tsv", ".parquet", ".parq", ".arrow", ".feather", ".xlsx", ".xls", ".orc", ".avro"}
JSON_EXTENSIONS = {".json"}
YAML_EXTENSIONS = {".yaml", ".yml"}
XML_EXTENSIONS = {".xml"}
HTML_EXTENSIONS = {".html", ".htm"}
MARKDOWN_EXTENSIONS = {".md", ".markdown"}
LOG_EXTENSIONS = {".log"}
SCRIPT_EXTENSIONS = {".sh", ".bat"}
MAKEFILE_NAMES = {"makefile", "gnumakefile", "bsdmakefile"}
INI_EXTENSIONS = {".ini", ".cfg"}
TOML_EXTENSIONS = {".toml"}
ENV_EXTENSIONS = {".env"}
NDJSON_EXTENSIONS = {".ndjson", ".jsonl"}
SQLITE_EXTENSIONS = {".sqlite", ".db"}
DUCKDB_EXTENSIONS = {".duckdb"}
FIREBIRD_EXTENSIONS = {".fdb", ".gdb", ".fbk"}
DBF_EXTENSIONS = {".dbf"}
DBF_SIDECAR_EXTENSIONS = {".dbt", ".fpt"}
PDF_EXTENSIONS = {".pdf"}
IMAGE_EXTENSIONS = {".gif", ".png", ".jpg", ".jpeg", ".bmp", ".webp", ".tif", ".tiff"}
PICKLE_EXTENSIONS = {".pkl", ".pickle"}
GZIP_EXTENSION = ".gz"
MAX_TABLE_ROWS = 2000
MAX_TABLE_COLS = 200
MAX_RENDER_ROWS = 500
MAX_RENDER_COLS = 50
MAX_TABULAR_PREVIEW_ROWS = 5000
MAX_EXCEL_RENDER_ROWS = 200
MAX_EXCEL_RENDER_COLS = 30
MAX_EXCEL_SCAN_ROWS = 1000
MAX_EXCEL_SCAN_COLS = 200
MAX_CELL_DISPLAY_LEN = 80
MAX_TEXT_PREVIEW_LINE_LEN = 240
MAX_PDF_PAGES = 20
MAX_PDF_CHARS = 120000
MAX_LOG_PREVIEW_LINES = 2000
MAX_LOG_PREVIEW_CHARS = 200000
CONVERTIBLE_EXTENSIONS = {".csv", ".tsv", ".parquet", ".parq", ".arrow", ".feather", ".orc", ".avro", ".json", ".pkl", ".pickle"}
ALLOW_UNSAFE_PICKLE_ENV = "FILES_VIEWER_ALLOW_PICKLE"


def configure_tzdir() -> None:
    # PyArrow may require an IANA tz database path for timezone-aware columns.
    if os.environ.get("TZDIR"):
        return
    try:
        zoneinfo_path = importlib_resources.files("tzdata").joinpath("zoneinfo")
    except Exception:
        return
    if zoneinfo_path.is_dir():
        os.environ["TZDIR"] = str(zoneinfo_path)


configure_tzdir()


class CellEditScreen(ModalScreen[str | None]):
    def __init__(self, current: str) -> None:
        super().__init__()
        self.current = current

    def compose(self) -> ComposeResult:
        with Vertical(id="edit_dialog"):
            yield Label("Edit cell value")
            yield Input(value=self.current, id="cell_input")
            yield Static("Enter: save, Esc: cancel", id="hint")

    @on(Input.Submitted, "#cell_input")
    def submit(self, event: Input.Submitted) -> None:
        self.dismiss(event.value)

    def key_escape(self) -> None:
        self.dismiss(None)


class ConvertScreen(ModalScreen[str | None]):
    def __init__(self, suggested: str) -> None:
        super().__init__()
        self.suggested = suggested

    def compose(self) -> ComposeResult:
        with Vertical(id="edit_dialog"):
            yield Label("Convert to file (e.g. output.parquet)")
            yield Input(value=self.suggested, id="convert_input")
            yield Static("Enter: convert, Esc: cancel", id="hint")

    @on(Input.Submitted, "#convert_input")
    def submit(self, event: Input.Submitted) -> None:
        self.dismiss(event.value.strip() or None)

    def key_escape(self) -> None:
        self.dismiss(None)


class NumberInputScreen(ModalScreen[int | None]):
    def __init__(self, title: str, hint: str, initial: str = "") -> None:
        super().__init__()
        self.title = title
        self.hint = hint
        self.initial = initial

    def compose(self) -> ComposeResult:
        with Vertical(id="edit_dialog"):
            yield Label(self.title)
            yield Input(value=self.initial, id="number_input")
            yield Static(self.hint, id="hint")

    @on(Input.Submitted, "#number_input")
    def submit(self, event: Input.Submitted) -> None:
        raw = event.value.strip()
        if not raw:
            self.dismiss(None)
            return
        try:
            self.dismiss(int(raw))
        except ValueError:
            self.dismiss(None)

    def key_escape(self) -> None:
        self.dismiss(None)


class TextInputScreen(ModalScreen[str | None]):
    def __init__(self, title: str, hint: str, initial: str = "") -> None:
        super().__init__()
        self.title = title
        self.hint = hint
        self.initial = initial

    def compose(self) -> ComposeResult:
        with Vertical(id="edit_dialog"):
            yield Label(self.title)
            yield Input(value=self.initial, id="text_input")
            yield Static(self.hint, id="hint")

    @on(Input.Submitted, "#text_input")
    def submit(self, event: Input.Submitted) -> None:
        self.dismiss(event.value.strip() or None)

    def key_escape(self) -> None:
        self.dismiss(None)


class HelpScreen(ModalScreen[None]):
    def compose(self) -> ComposeResult:
        with Vertical(id="help_dialog"):
            yield Label("Help")
            yield Static(
                "\n".join(
                    [
                        "Files Viewer TUI: browse, preview, edit and convert data/text files.",
                        "Types: csv/tsv, parquet/parq, orc, avro, arrow/feather, xlsx/xls, sqlite/db/duckdb/firebird, dbf,",
                        "json/ndjson/jsonl/yaml/yml/xml/html/markdown/log/sh/bat/makefile/ini/cfg/toml/env,",
                        "plus gzip-compressed text/tabular variants (e.g. csv.gz, tsv.gz, json.gz, log.gz).",
                        "pdf, and images (gif/png/jpg/jpeg/bmp/webp/tif/tiff).",
                        "",
                        "Ctrl+O  Focus Tree",
                        "Backspace  Parent Dir",
                        "Alt+Left / Alt+Right  Excel sheets",
                        "Ctrl+R  Open HTML in browser",
                        "Ctrl+T  Convert table format",
                        "Alt+P / F6  Promote lazy table to full mode",
                        "Ctrl+S  Save",
                        "/  Search in text view",
                        "F3 / Shift+F3  Next / previous match",
                        "Ctrl+H  Toggle search highlight",
                        "Alt+S / F7  Sort selected column",
                        "Ctrl+K  Visible columns",
                        "g  Jump to table page",
                        "[ / ]  Decrease / increase page size",
                        "e  Edit selected cell",
                        "?  Open this help",
                        "Esc / q  Close or quit",
                    ]
                )
            )
            yield Static("Press Esc to close", id="hint")

    def key_escape(self) -> None:
        self.dismiss(None)


class ConfirmPickleScreen(ModalScreen[bool]):
    def compose(self) -> ComposeResult:
        with Vertical(id="pickle_dialog"):
            yield Label("Unsafe pickle load")
            yield Static("Pickle files may execute arbitrary code.\nLoad only if the file is trusted.")
            with Horizontal(id="pickle_buttons"):
                yield Button("Cancel", id="pickle_cancel")
                yield Button("Load Trusted Pickle", id="pickle_confirm")

    @on(Button.Pressed, "#pickle_cancel")
    def cancel(self) -> None:
        self.dismiss(False)

    @on(Button.Pressed, "#pickle_confirm")
    def confirm(self) -> None:
        self.dismiss(True)

    def key_escape(self) -> None:
        self.dismiss(False)


class FileLoaded(Message):
    def __init__(self, path: Path) -> None:
        self.path = path
        super().__init__()


class ColoredDirectoryTree(DirectoryTree):
    EXTENSION_COLORS = {
        ".csv": "#6ee7b7",
        ".tsv": "#6ee7b7",
        ".parquet": "#fca5a5",
        ".parq": "#fca5a5",
        ".orc": "#fda4af",
        ".avro": "#fda4af",
        ".arrow": "#93c5fd",
        ".feather": "#93c5fd",
        ".xlsx": "#86efac",
        ".xls": "#86efac",
        ".json": "#fcd34d",
        ".ndjson": "#fcd34d",
        ".jsonl": "#fcd34d",
        ".yaml": "#fcd34d",
        ".yml": "#fcd34d",
        ".ini": "#fbbf24",
        ".cfg": "#fbbf24",
        ".toml": "#f59e0b",
        ".env": "#84cc16",
        ".sqlite": "#fca5a5",
        ".db": "#fca5a5",
        ".duckdb": "#fca5a5",
        ".fdb": "#fca5a5",
        ".gdb": "#fca5a5",
        ".fbk": "#fca5a5",
        ".dbf": "#fca5a5",
        ".dbt": "#94a3b8",
        ".fpt": "#94a3b8",
        ".gz": "#cbd5e1",
        ".log": "#9ca3af",
        ".sh": "#94a3b8",
        ".bat": "#94a3b8",
        ".mk": "#94a3b8",
        ".xml": "#fdba74",
        ".html": "#c4b5fd",
        ".htm": "#c4b5fd",
        ".pdf": "#f87171",
        ".gif": "#67e8f9",
        ".png": "#67e8f9",
        ".jpg": "#67e8f9",
        ".jpeg": "#67e8f9",
        ".bmp": "#67e8f9",
        ".webp": "#67e8f9",
        ".tif": "#67e8f9",
        ".tiff": "#67e8f9",
        ".pkl": "#a78bfa",
        ".pickle": "#a78bfa",
    }

    @staticmethod
    def _human_size(num_bytes: int) -> str:
        units = ["B", "KB", "MB", "GB", "TB"]
        size = float(num_bytes)
        for unit in units:
            if size < 1024 or unit == units[-1]:
                if unit == "B":
                    return f"{int(size)}{unit}"
                return f"{size:.1f}{unit}"
            size /= 1024
        return f"{int(num_bytes)}B"

    @staticmethod
    def _fixed(text: str, width: int, align: str = "left") -> str:
        if len(text) <= width:
            return text.rjust(width) if align == "right" else text.ljust(width)
        if width <= 1:
            return text[:width]
        return text[: width - 1] + "…"

    def render_label(self, node: TreeNode, base_style, style):
        label = Text()
        try:
            path = Path(str(node.data.path))
            size_text = "-"
            date_text = "-"
            if path.exists():
                stat = path.stat()
                if path.is_file():
                    size_text = self._human_size(stat.st_size)
                date_text = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M")

            name_text = path.name or str(path)
            fixed_name = self._fixed(name_text, 35)
            fixed_size = self._fixed(size_text, 8, align="right")
            fixed_date = self._fixed(date_text, 16)
            row_text = f"{fixed_name} {fixed_size} {fixed_date}"

            if path.is_file():
                color = self.EXTENSION_COLORS.get(path.suffix.lower())
                if color:
                    label.append(Text(row_text, style=color))
                else:
                    label.append(Text(row_text, style="#9cc4ff"))
            else:
                label.append(Text(row_text, style="#ffd166"))
        except Exception:
            label = super().render_label(node, base_style, style)
        return label


class FilesViewerApp(App[None]):
    CSS = """
    Screen {
        layout: vertical;
    }

    #main {
        height: 1fr;
    }

    #tree {
        width: 34%;
        border: solid $primary;
        background: #0b1220;
        color: #dbe7ff;
    }

    #tree .tree--guides {
        color: #35507a;
    }

    #tree .tree--label {
        color: #c9ddff;
    }

    #tree .tree--cursor {
        background: #1f3a5f;
        color: #ffffff;
        text-style: bold;
    }

    #tree .directory-tree--folder {
        color: #ffd166;
        text-style: bold;
    }

    #tree .directory-tree--file {
        color: #9cc4ff;
    }

    #content {
        width: 66%;
        border: solid $secondary;
        overflow: auto auto;
    }

    #text_tab {
        overflow: auto auto;
    }

    #text {
        height: 1fr;
        overflow: auto auto;
    }

    #image_tab {
        overflow: scroll scroll;
    }

    #image_preview {
        height: 1fr;
        overflow: scroll scroll;
        padding: 1;
    }

    #markdown_scroll {
        height: 1fr;
        overflow: auto auto;
    }

    #markdown_preview {
        padding: 1;
    }

    #json_tree_tab {
        overflow: auto auto;
    }

    #json_tree {
        height: 1fr;
    }

    #status {
        height: 1;
        padding: 0 1;
        content-align: left middle;
    }

    #context_shortcuts {
        height: 1;
        padding: 0 1;
        color: #9fb3d8;
        content-align: left middle;
    }

    #pager {
        height: 3;
        padding: 0 1;
        content-align: left middle;
    }

    #page_first, #page_prev, #page_next, #page_last {
        width: 6;
        min-width: 6;
        max-width: 6;
        padding: 0 0;
        margin: 0 1;
        content-align: center middle;
    }

    #page_info {
        width: 1fr;
        content-align: center middle;
    }

    #table_filter_bar {
        height: 3;
        padding: 0 1;
        content-align: left middle;
    }

    #table_filter_input {
        width: 1fr;
    }

    #edit_dialog {
        width: 60;
        height: 7;
        border: round $warning;
        background: $panel;
        padding: 1;
        align: center middle;
    }

    #help_dialog {
        width: 110;
        height: 24;
        border: round $warning;
        background: $panel;
        padding: 1;
        align: center middle;
    }

    #pickle_dialog {
        width: 76;
        height: 11;
        border: round $error;
        background: $panel;
        padding: 1;
        align: center middle;
    }

    #pickle_buttons {
        height: auto;
        padding-top: 1;
    }
    """

    BINDINGS = [
        Binding("ctrl+o", "focus_tree", "Focus Tree"),
        Binding("backspace", "go_parent_dir", "Parent Dir"),
        Binding("alt+left", "prev_sheet", "Prev Sheet", show=False),
        Binding("alt+right", "next_sheet", "Next Sheet", show=False),
        Binding("ctrl+r", "render_html", "Render HTML", show=False),
        Binding("ctrl+t", "convert_file", "Convert", show=False),
        Binding("alt+p", "promote_full_table", "Promote Full", show=False),
        Binding("f6", "promote_full_table", "Promote Full", show=False),
        Binding("ctrl+s", "save", "Save"),
        Binding("slash", "search_text", "Search", show=False),
        Binding("f3", "next_match", "Next Match", show=False),
        Binding("shift+f3", "prev_match", "Prev Match", show=False),
        Binding("ctrl+h", "toggle_search_highlight", "Highlight"),
        Binding("alt+s", "sort_selected_column", "Sort Column", show=False),
        Binding("f7", "sort_selected_column", "Sort Column", show=False),
        Binding("ctrl+k", "pick_visible_columns", "Visible Cols", show=False),
        Binding("g", "jump_to_page", "Jump Page", show=False),
        Binding("left_square_bracket", "decrease_page_size", "Pg -", show=False),
        Binding("right_square_bracket", "increase_page_size", "Pg +", show=False),
        Binding("e", "edit_cell", "Edit Cell", show=False),
        Binding("question_mark", "show_help", "Help"),
        Binding("q", "quit", "Quit"),
    ]

    def __init__(self) -> None:
        super().__init__()
        self.tree_root = Path.cwd().resolve()
        self.current_path: Path | None = None
        self.current_kind: str | None = None
        self.current_df: pd.DataFrame | None = None
        self.current_df_source: pd.DataFrame | None = None
        self.lazy_delimited_context: dict[str, Any] | None = None
        self.visible_columns: list[str] | None = None
        self.current_sort: tuple[str, bool] | None = None  # (column, ascending)
        self.current_table_preview_only = False
        self.current_excel_sheet: str | int = 0
        self.current_excel_sheets: list[str] = []
        self.current_excel_sheet_index = 0
        self.table_page = 0
        self.table_page_size = 32
        self.allow_unsafe_pickle_session = False
        self.current_log_text = ""
        self.log_text_colored_ready = False
        self.table_filter_query = ""
        self.last_text_search = ""
        self.text_search_matches: list[int] = []
        self.text_search_index = -1
        self.text_search_highlight_all = False
        self.preview_cache: dict[tuple[str, int, str], dict[str, Any]] = {}
        self.cache_hits = 0
        self.last_load_ms = 0
        self.is_page_loading = False

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        with Horizontal(id="main"):
            yield ColoredDirectoryTree(str(self.tree_root), id="tree")
            with Vertical(id="content"):
                with TabbedContent(id="viewer_tabs"):
                    with TabPane("DB Schema", id="db_schema_tab"):
                        yield Tree("DB Schema", id="db_schema_tree")
                    with TabPane("Table", id="table_tab"):
                        with Horizontal(id="table_filter_bar"):
                            yield Input(placeholder="Filter (global or col=value, col==value, col!=value, col~regex) + Enter", id="table_filter_input")
                            yield Button("Clear", id="table_filter_clear")
                        yield DataTable(id="table")
                    with TabPane("Schema", id="schema_tab"):
                        yield DataTable(id="schema_table")
                    with TabPane("Text", id="text_tab"):
                        yield Static("", id="log_text_preview")
                        yield TextArea(id="text")
                    with TabPane("JSON Tree", id="json_tree_tab"):
                        yield Tree("JSON", id="json_tree")
                    with TabPane("Markdown", id="markdown_tab"):
                        with VerticalScroll(id="markdown_scroll"):
                            yield Static("", id="markdown_preview")
                    with TabPane("Image", id="image_tab"):
                        yield Static("", id="image_preview")
                with Horizontal(id="pager"):
                    yield Button("<<", id="page_first")
                    yield Button("<", id="page_prev")
                    yield Static("Page 1/1", id="page_info")
                    yield Button(">", id="page_next")
                    yield Button(">>", id="page_last")
        yield Static("Open a file from the tree.", id="status")
        yield Static("", id="context_shortcuts")
        yield Footer()

    def on_mount(self) -> None:
        table = self.query_one(DataTable)
        schema_table = self.query_one("#schema_table", DataTable)
        text = self.query_one(TextArea)
        filter_input = self.query_one("#table_filter_input", Input)
        table.show_cursor = True
        schema_table.show_cursor = False
        text.show_line_numbers = True
        text.soft_wrap = False
        filter_input.value = ""
        self.query_one("#log_text_preview", Static).display = False
        self.query_one("#pager", Horizontal).display = False
        self.update_context_shortcuts()

    @on(DirectoryTree.FileSelected)
    def file_selected(self, event: DirectoryTree.FileSelected) -> None:
        self.open_file(Path(event.path))

    def open_file(self, path: Path) -> None:
        if path.suffix.lower() in PICKLE_EXTENSIONS and not self.is_pickle_loading_allowed():
            self.push_screen(ConfirmPickleScreen(), lambda ok: self.on_pickle_confirm(path, ok))
            return
        self.clear_viewers()
        self.query_one("#status", Static).update(f"Loading: {path}")
        self.show_text(f"Loading file...\n{path}")
        self.load_file_worker(path)

    def on_pickle_confirm(self, path: Path, confirmed: bool) -> None:
        if not confirmed:
            self.query_one("#status", Static).update(f"Cancelled unsafe file load: {path}")
            self.show_text("Pickle load cancelled.")
            return
        self.allow_unsafe_pickle_session = True
        self.open_file(path)

    def clear_viewers(self) -> None:
        table = self.query_one(DataTable)
        schema_table = self.query_one("#schema_table", DataTable)
        text = self.query_one(TextArea)
        filter_input = self.query_one("#table_filter_input", Input)
        log_preview = self.query_one("#log_text_preview", Static)
        image = self.query_one("#image_preview", Static)
        markdown = self.query_one("#markdown_preview", Static)
        json_tree = self.query_one("#json_tree", Tree)
        db_schema_tree = self.query_one("#db_schema_tree", Tree)
        table.clear(columns=True)
        schema_table.clear(columns=True)
        text.text = ""
        log_preview.update("")
        log_preview.display = False
        text.display = True
        image.update("")
        markdown.update("")
        json_tree.clear()
        json_tree.root.set_label("JSON")
        db_schema_tree.clear()
        db_schema_tree.root.set_label("DB Schema")
        self.query_one("#pager", Horizontal).display = False
        self.current_log_text = ""
        self.log_text_colored_ready = False
        self.current_df = None
        self.current_df_source = None
        self.lazy_delimited_context = None
        self.visible_columns = None
        self.current_sort = None
        self.table_filter_query = ""
        filter_input.value = ""

    @work(thread=True)
    def load_file_worker(self, path: Path) -> None:
        suffix = path.suffix.lower()
        try:
            cache_key = self.make_cache_key(path)
            cached = self.preview_cache.get(cache_key)
            if cached is not None:
                self.cache_hits += 1
                self.call_from_thread(self.apply_loaded_result, self.clone_cached_result(cached))
                return
            started = datetime.now()
            if suffix in {".xlsx", ".xls"}:
                result = self.prepare_excel_text_result(path)
            elif suffix in IMAGE_EXTENSIONS:
                result = self.prepare_image_result(path)
            elif suffix == GZIP_EXTENSION:
                result = self.prepare_gzip_result(path)
            elif suffix in SQLITE_EXTENSIONS:
                result = self.prepare_sqlite_result(path)
            elif suffix in DUCKDB_EXTENSIONS:
                result = self.prepare_duckdb_result(path)
            elif suffix in FIREBIRD_EXTENSIONS:
                result = self.prepare_firebird_result(path)
            elif suffix in DBF_EXTENSIONS:
                result = self.prepare_dbf_result(path)
            elif suffix in DBF_SIDECAR_EXTENSIONS:
                result = self.prepare_dbf_sidecar_result(path)
            elif suffix in TABULAR_EXTENSIONS:
                result = self.prepare_table_result(path)
            elif suffix in PICKLE_EXTENSIONS:
                result = self.prepare_pickle_result(path)
            elif suffix in JSON_EXTENSIONS:
                raw_text = self.read_text_with_fallback(path)
                data = json.loads(raw_text)
                result = {
                    "path": path,
                    "kind": "json",
                    "text": json.dumps(data, indent=2, ensure_ascii=False),
                    "json_obj": data,
                }
            elif suffix in NDJSON_EXTENSIONS:
                raw_text = self.read_text_with_fallback(path)
                data = self.read_ndjson_preview(raw_text)
                result = {
                    "path": path,
                    "kind": "json",
                    "text": json.dumps(data, indent=2, ensure_ascii=False),
                    "json_obj": data,
                }
            elif suffix in XML_EXTENSIONS:
                xml_text = self.read_xml_with_fallback(path)
                result = {
                    "path": path,
                    "kind": "text",
                    "text": xml_text,
                    "json_obj": self.xml_to_json_obj(xml_text),
                }
            elif suffix in HTML_EXTENSIONS:
                result = {"path": path, "kind": "text", "text": self.read_html_with_fallback(path)}
            elif suffix in MARKDOWN_EXTENSIONS:
                result = {"path": path, "kind": "text", "text": self.read_text_with_fallback(path)}
            elif suffix in LOG_EXTENSIONS:
                log_text = self.read_text_with_fallback(path)
                preview_text, truncated = self.limit_log_preview(log_text)
                result = {
                    "path": path,
                    "kind": "log",
                    "text": preview_text,
                    "truncated": truncated,
                }
            elif suffix in PDF_EXTENSIONS:
                result = {"path": path, "kind": "text", "text": self.read_pdf_preview(path)}
            else:
                result = {"path": path, "kind": "text", "text": self.read_text_with_fallback(path)}
            self.preview_cache[cache_key] = result
            self.last_load_ms = int((datetime.now() - started).total_seconds() * 1000)
            self.call_from_thread(self.apply_loaded_result, result)
        except Exception as exc:
            self.call_from_thread(self.handle_load_error, path, exc)

    def make_cache_key(self, path: Path) -> tuple[str, int, str]:
        stat = path.stat()
        return (str(path.resolve()), int(stat.st_mtime_ns), path.suffix.lower())

    def clone_cached_result(self, result: dict[str, Any]) -> dict[str, Any]:
        cloned = copy.deepcopy(result)
        df = cloned.get("df")
        if isinstance(df, pd.DataFrame):
            cloned["df"] = df.copy()
        return cloned

    def prepare_table_result(self, path: Path) -> dict[str, Any]:
        suffix = path.suffix.lower()
        preview_only = False
        excel_sheets: list[str] = []
        excel_sheet_index = 0
        excel_sheet: str | int = 0

        if suffix == ".csv":
            lazy = self.prepare_lazy_delimited_context(path, sep=",")
            return {
                "path": path,
                "kind": "table",
                "df": lazy["df"],
                "preview_only": preview_only,
                "excel_sheets": excel_sheets,
                "excel_sheet_index": excel_sheet_index,
                "excel_sheet": excel_sheet,
                "lazy_delimited": lazy,
            }
        elif suffix == ".tsv":
            lazy = self.prepare_lazy_delimited_context(path, sep="\t")
            return {
                "path": path,
                "kind": "table",
                "df": lazy["df"],
                "preview_only": preview_only,
                "excel_sheets": excel_sheets,
                "excel_sheet_index": excel_sheet_index,
                "excel_sheet": excel_sheet,
                "lazy_delimited": lazy,
            }
        elif suffix in {".parquet", ".parq"}:
            df = self.read_arrow_preview(path, format_name="parquet")
        elif suffix == ".orc":
            df = self.read_arrow_preview(path, format_name="orc")
        elif suffix == ".avro":
            df = self.read_avro_preview(path)
        elif suffix in {".xlsx", ".xls"}:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            excel_sheets = list(workbook.sheetnames)
            workbook.close()
            excel_sheet = excel_sheets[0] if excel_sheets else 0
            df = self.read_excel_sheet(path, excel_sheet)
        else:
            df = self.read_arrow_preview(path, format_name="ipc")

        return {
            "path": path,
            "kind": "table",
            "df": df,
            "preview_only": preview_only,
            "excel_sheets": excel_sheets,
            "excel_sheet_index": excel_sheet_index,
            "excel_sheet": excel_sheet,
        }

    def prepare_gzip_result(self, path: Path) -> dict[str, Any]:
        inner_suffix = self.get_gzip_inner_suffix(path)
        if inner_suffix == ".csv":
            df = self.read_gzip_delimited_with_fallback(path, sep=",")
            return {
                "path": path,
                "kind": "table",
                "df": df,
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
            }
        if inner_suffix == ".tsv":
            df = self.read_gzip_delimited_with_fallback(path, sep="\t")
            return {
                "path": path,
                "kind": "table",
                "df": df,
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
            }
        if inner_suffix in JSON_EXTENSIONS:
            raw_text = self.read_gzip_text_with_fallback(path)
            data = json.loads(raw_text)
            return {
                "path": path,
                "kind": "json",
                "text": json.dumps(data, indent=2, ensure_ascii=False),
                "json_obj": data,
            }
        if inner_suffix in NDJSON_EXTENSIONS:
            raw_text = self.read_gzip_text_with_fallback(path)
            data = self.read_ndjson_preview(raw_text)
            return {
                "path": path,
                "kind": "json",
                "text": json.dumps(data, indent=2, ensure_ascii=False),
                "json_obj": data,
            }

        text = self.read_gzip_text_with_fallback(path)
        return {"path": path, "kind": "text", "text": text}

    def get_gzip_inner_suffix(self, path: Path) -> str:
        suffixes = [s.lower() for s in path.suffixes]
        if len(suffixes) >= 2 and suffixes[-1] == GZIP_EXTENSION:
            return suffixes[-2]
        return ""

    def prepare_sqlite_result(self, path: Path) -> dict[str, Any]:
        conn = sqlite3.connect(str(path))
        try:
            tables_df = pd.read_sql_query(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name",
                conn,
            )
            table_names = tables_df["name"].tolist() if not tables_df.empty else []
            if not table_names:
                return {
                    "path": path,
                    "kind": "text",
                    "text": f"SQLite DB: {path.name}\n\nNo user tables found.",
                }
            first_table = table_names[0]
            preview_sql = f'SELECT * FROM "{first_table}" LIMIT {MAX_RENDER_ROWS}'
            df = pd.read_sql_query(preview_sql, conn)
            lines = [
                f"SQLite DB: {path.name}",
                f"Tables: {len(table_names)}",
                f"Preview table: {first_table}",
                "",
                "Table list:",
            ] + [f"- {name}" for name in table_names[:100]]
            if len(table_names) > 100:
                lines.append("...")
            return {
                "path": path,
                "kind": "table",
                "df": df,
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
                "text": "\n".join(lines),
                "db_schema": self.read_sqlite_schema(conn, table_names),
            }
        finally:
            conn.close()

    def prepare_duckdb_result(self, path: Path) -> dict[str, Any]:
        try:
            import duckdb  # type: ignore
        except Exception as exc:
            raise RuntimeError("duckdb package is required to open .duckdb files") from exc

        conn = duckdb.connect(str(path), read_only=True)
        try:
            tables_df = conn.execute(
                """
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
                  AND table_type = 'BASE TABLE'
                ORDER BY table_name
                """
            ).fetchdf()
            table_names = tables_df["table_name"].tolist() if not tables_df.empty else []
            if not table_names:
                return {
                    "path": path,
                    "kind": "text",
                    "text": f"DuckDB: {path.name}\n\nNo user tables found.",
                }
            first_table = table_names[0]
            preview_sql = f'SELECT * FROM "{first_table}" LIMIT {MAX_RENDER_ROWS}'
            df = conn.execute(preview_sql).fetchdf()
            lines = [
                f"DuckDB: {path.name}",
                f"Tables: {len(table_names)}",
                f"Preview table: {first_table}",
                "",
                "Table list:",
            ] + [f"- {name}" for name in table_names[:100]]
            if len(table_names) > 100:
                lines.append("...")
            return {
                "path": path,
                "kind": "table",
                "df": df,
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
                "text": "\n".join(lines),
                "db_schema": self.read_duckdb_schema(conn),
            }
        finally:
            conn.close()

    def prepare_firebird_result(self, path: Path) -> dict[str, Any]:
        conn = self.connect_firebird(path)
        try:
            table_rows = self.firebird_query(
                conn,
                """
                SELECT TRIM(r.RDB$RELATION_NAME) AS NAME
                FROM RDB$RELATIONS r
                WHERE COALESCE(r.RDB$SYSTEM_FLAG, 0) = 0
                  AND COALESCE(r.RDB$VIEW_BLR, '') = ''
                ORDER BY 1
                """,
            )
            table_names = [str(r.get("NAME", "")).strip() for r in table_rows if str(r.get("NAME", "")).strip()]
            if not table_names:
                return {
                    "path": path,
                    "kind": "text",
                    "text": f"Firebird DB: {path.name}\n\nNo user tables found.",
                }
            first_table = table_names[0]
            preview_sql = f'SELECT * FROM "{first_table}" ROWS {MAX_RENDER_ROWS}'
            df = self.firebird_query_df(conn, preview_sql)
            lines = [
                f"Firebird DB: {path.name}",
                f"Tables: {len(table_names)}",
                f"Preview table: {first_table}",
                "",
                "Table list:",
            ] + [f"- {name}" for name in table_names[:100]]
            if len(table_names) > 100:
                lines.append("...")
            return {
                "path": path,
                "kind": "table",
                "df": df,
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
                "text": "\n".join(lines),
                "db_schema": self.read_firebird_schema(conn),
            }
        finally:
            self.close_firebird(conn)

    def prepare_dbf_result(self, path: Path) -> dict[str, Any]:
        try:
            from dbfread import DBF  # type: ignore
        except Exception as exc:
            raise RuntimeError("dbfread package is required to open .dbf files") from exc

        try:
            table = DBF(str(path), load=False, char_decode_errors="ignore")
            field_names = list(table.field_names or [])
            rows: list[dict[str, Any]] = []
            for i, rec in enumerate(table):
                if i >= MAX_RENDER_ROWS:
                    break
                rows.append(dict(rec))
            df = pd.DataFrame(rows, columns=field_names if field_names else None)
            info = [
                f"DBF: {path.name}",
                f"Fields: {len(field_names)}",
                f"Preview rows: {len(df)}",
            ]
            return {
                "path": path,
                "kind": "table",
                "df": df,
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
                "text": "\n".join(info),
            }
        except Exception as exc:
            raise RuntimeError(f"Failed to read DBF: {exc}") from exc

    def prepare_dbf_sidecar_result(self, path: Path) -> dict[str, Any]:
        stem = path.stem
        sibling_dbf = path.with_suffix(".dbf")
        lines = [
            f"DBF sidecar file: {path.name}",
            "",
            "This file stores memo/index companion data and is not directly previewed as a table.",
            f"Open the corresponding DBF table file instead: {sibling_dbf.name}",
        ]
        return {"path": path, "kind": "text", "text": "\n".join(lines)}

    def read_sqlite_schema(self, conn: sqlite3.Connection, table_names: list[str]) -> dict[str, Any]:
        schema: dict[str, Any] = {"tables": {}, "views": [], "indexes": [], "triggers": []}
        for table_name in table_names:
            cols = pd.read_sql_query(f'PRAGMA table_info("{table_name}")', conn).to_dict(orient="records")
            fks = pd.read_sql_query(f'PRAGMA foreign_key_list("{table_name}")', conn).to_dict(orient="records")
            idx_list = pd.read_sql_query(f'PRAGMA index_list("{table_name}")', conn).to_dict(orient="records")
            indexes: list[dict[str, Any]] = []
            for idx in idx_list:
                idx_name = str(idx.get("name", ""))
                idx_cols = pd.read_sql_query(f'PRAGMA index_info("{idx_name}")', conn).to_dict(orient="records")
                indexes.append(
                    {
                        "name": idx_name,
                        "unique": bool(idx.get("unique", 0)),
                        "origin": idx.get("origin"),
                        "columns": [c.get("name") for c in idx_cols if c.get("name") is not None],
                    }
                )
            schema["tables"][table_name] = {
                "columns": cols,
                "foreign_keys": fks,
                "indexes": indexes,
            }
        views_df = pd.read_sql_query(
            "SELECT name, sql FROM sqlite_master WHERE type='view' AND name NOT LIKE 'sqlite_%' ORDER BY name",
            conn,
        )
        schema["views"] = views_df.to_dict(orient="records") if not views_df.empty else []

        indexes_df = pd.read_sql_query(
            "SELECT name, tbl_name, sql FROM sqlite_master WHERE type='index' AND name NOT LIKE 'sqlite_%' ORDER BY tbl_name, name",
            conn,
        )
        global_indexes: list[dict[str, Any]] = []
        for rec in indexes_df.to_dict(orient="records") if not indexes_df.empty else []:
            idx_name = str(rec.get("name", ""))
            idx_cols = pd.read_sql_query(f'PRAGMA index_info("{idx_name}")', conn).to_dict(orient="records")
            global_indexes.append(
                {
                    "name": idx_name,
                    "table": rec.get("tbl_name"),
                    "sql": rec.get("sql"),
                    "columns": [c.get("name") for c in idx_cols if c.get("name") is not None],
                }
            )
        schema["indexes"] = global_indexes

        triggers_df = pd.read_sql_query(
            "SELECT name, tbl_name, sql FROM sqlite_master WHERE type='trigger' ORDER BY tbl_name, name",
            conn,
        )
        schema["triggers"] = triggers_df.to_dict(orient="records") if not triggers_df.empty else []
        return schema

    def read_duckdb_schema(self, conn: Any) -> dict[str, Any]:
        schema: dict[str, Any] = {"tables": {}, "views": [], "indexes": [], "triggers": []}

        tables_df = conn.execute(
            """
            SELECT table_schema, table_name
            FROM information_schema.tables
            WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
              AND table_type = 'BASE TABLE'
            ORDER BY table_schema, table_name
            """
        ).fetchdf()
        views_df = conn.execute(
            """
            SELECT table_schema, table_name
            FROM information_schema.tables
            WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
              AND table_type = 'VIEW'
            ORDER BY table_schema, table_name
            """
        ).fetchdf()
        cols_df = conn.execute(
            """
            SELECT table_schema, table_name, column_name, data_type, is_nullable
            FROM information_schema.columns
            WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
            ORDER BY table_schema, table_name, ordinal_position
            """
        ).fetchdf()

        table_keys = []
        if not tables_df.empty:
            for _, row in tables_df.iterrows():
                schema_name = str(row["table_schema"])
                table_name = str(row["table_name"])
                key = f"{schema_name}.{table_name}"
                table_keys.append((schema_name, table_name, key))

        for schema_name, table_name, key in table_keys:
            subset = cols_df[(cols_df["table_schema"] == schema_name) & (cols_df["table_name"] == table_name)]
            columns: list[dict[str, Any]] = []
            for _, c in subset.iterrows():
                columns.append(
                    {
                        "name": str(c["column_name"]),
                        "type": str(c["data_type"]),
                        "notnull": 0 if str(c["is_nullable"]).upper() == "YES" else 1,
                        "pk": 0,
                        "dflt_value": None,
                    }
                )
            schema["tables"][key] = {"columns": columns, "foreign_keys": [], "indexes": []}

        if not views_df.empty:
            schema["views"] = [
                {"name": f'{str(r["table_schema"])}.{str(r["table_name"])}', "sql": None}
                for _, r in views_df.iterrows()
            ]
        return schema

    def read_firebird_schema(self, conn: Any) -> dict[str, Any]:
        schema: dict[str, Any] = {"tables": {}, "views": [], "indexes": [], "triggers": []}

        table_rows = self.firebird_query(
            conn,
            """
            SELECT TRIM(r.RDB$RELATION_NAME) AS NAME
            FROM RDB$RELATIONS r
            WHERE COALESCE(r.RDB$SYSTEM_FLAG, 0) = 0
              AND COALESCE(r.RDB$VIEW_BLR, '') = ''
            ORDER BY 1
            """,
        )
        view_rows = self.firebird_query(
            conn,
            """
            SELECT TRIM(r.RDB$RELATION_NAME) AS NAME
            FROM RDB$RELATIONS r
            WHERE COALESCE(r.RDB$SYSTEM_FLAG, 0) = 0
              AND r.RDB$VIEW_BLR IS NOT NULL
            ORDER BY 1
            """,
        )
        index_rows = self.firebird_query(
            conn,
            """
            SELECT TRIM(i.RDB$INDEX_NAME) AS NAME,
                   TRIM(i.RDB$RELATION_NAME) AS TABLE_NAME,
                   i.RDB$UNIQUE_FLAG AS UNIQUE_FLAG
            FROM RDB$INDICES i
            WHERE COALESCE(i.RDB$SYSTEM_FLAG, 0) = 0
            ORDER BY 2, 1
            """,
        )
        trigger_rows = self.firebird_query(
            conn,
            """
            SELECT TRIM(t.RDB$TRIGGER_NAME) AS NAME,
                   TRIM(t.RDB$RELATION_NAME) AS TABLE_NAME
            FROM RDB$TRIGGERS t
            WHERE COALESCE(t.RDB$SYSTEM_FLAG, 0) = 0
            ORDER BY 2, 1
            """,
        )

        for row in table_rows:
            tname = str(row.get("NAME", "")).strip()
            if not tname:
                continue
            cols = self.firebird_query(
                conn,
                f"""
                SELECT TRIM(rf.RDB$FIELD_NAME) AS NAME,
                       TRIM(f.RDB$FIELD_TYPE) AS FIELD_TYPE_CODE,
                       COALESCE(rf.RDB$NULL_FLAG, 0) AS NOTNULL_FLAG
                FROM RDB$RELATION_FIELDS rf
                JOIN RDB$FIELDS f ON f.RDB$FIELD_NAME = rf.RDB$FIELD_SOURCE
                WHERE rf.RDB$RELATION_NAME = '{tname}'
                ORDER BY rf.RDB$FIELD_POSITION
                """,
            )
            columns = []
            for c in cols:
                columns.append(
                    {
                        "name": str(c.get("NAME", "")).strip(),
                        "type": str(c.get("FIELD_TYPE_CODE", "")).strip(),
                        "notnull": int(c.get("NOTNULL_FLAG", 0)),
                        "pk": 0,
                        "dflt_value": None,
                    }
                )
            schema["tables"][tname] = {"columns": columns, "foreign_keys": [], "indexes": []}

        for v in view_rows:
            vname = str(v.get("NAME", "")).strip()
            if vname:
                schema["views"].append({"name": vname, "sql": None})

        for idx in index_rows:
            iname = str(idx.get("NAME", "")).strip()
            tname = str(idx.get("TABLE_NAME", "")).strip()
            unique = bool(int(idx.get("UNIQUE_FLAG") or 0))
            idx_cols = self.firebird_query(
                conn,
                f"""
                SELECT TRIM(s.RDB$FIELD_NAME) AS NAME
                FROM RDB$INDEX_SEGMENTS s
                WHERE s.RDB$INDEX_NAME = '{iname}'
                ORDER BY s.RDB$FIELD_POSITION
                """,
            )
            cols = [str(c.get("NAME", "")).strip() for c in idx_cols if str(c.get("NAME", "")).strip()]
            schema["indexes"].append({"name": iname, "table": tname, "sql": None, "columns": cols})
            if tname in schema["tables"]:
                schema["tables"][tname]["indexes"].append(
                    {"name": iname, "unique": unique, "origin": None, "columns": cols}
                )

        for trg in trigger_rows:
            tname = str(trg.get("TABLE_NAME", "")).strip()
            trname = str(trg.get("NAME", "")).strip()
            if trname:
                schema["triggers"].append({"name": trname, "tbl_name": tname, "sql": None})
        return schema

    def connect_firebird(self, path: Path) -> Any:
        user = os.environ.get("FIREBIRD_USER", "SYSDBA")
        password = os.environ.get("FIREBIRD_PASSWORD", "masterkey")
        host = os.environ.get("FIREBIRD_HOST", "localhost")

        # Try modern firebird-driver first
        try:
            from firebird.driver import connect  # type: ignore

            dsn = f"{host}:{path}"
            return connect(dsn=dsn, user=user, password=password)
        except Exception:
            pass

        # Fallback to legacy fdb driver
        try:
            import fdb  # type: ignore

            dsn = f"{host}:{path}"
            return fdb.connect(dsn=dsn, user=user, password=password)
        except Exception as exc:
            raise RuntimeError(
                "Firebird connection failed. Install `firebird-driver` or `fdb`, ensure Firebird server is reachable, "
                "and set FIREBIRD_USER/FIREBIRD_PASSWORD/FIREBIRD_HOST if needed."
            ) from exc

    def close_firebird(self, conn: Any) -> None:
        try:
            conn.close()
        except Exception:
            pass

    def firebird_query(self, conn: Any, sql: str) -> list[dict[str, Any]]:
        cur = conn.cursor()
        try:
            cur.execute(sql)
            rows = cur.fetchall()
            desc = cur.description or []
            cols = [str(d[0]).strip() for d in desc]
            result: list[dict[str, Any]] = []
            for row in rows:
                result.append({cols[i]: row[i] for i in range(len(cols))})
            return result
        finally:
            try:
                cur.close()
            except Exception:
                pass

    def firebird_query_df(self, conn: Any, sql: str) -> pd.DataFrame:
        rows = self.firebird_query(conn, sql)
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame(rows)

    def prepare_pickle_result(self, path: Path) -> dict[str, Any]:
        if not self.is_pickle_loading_allowed():
            return {
                "path": path,
                "kind": "text",
                "text": (
                    "Pickle preview is disabled for safety.\n\n"
                    "Reason: pickle loading can execute arbitrary code from untrusted files.\n"
                    "Use the in-app confirmation dialog to load trusted pickle files."
                ),
            }

        stream_df = self.read_pickle_stream_preview(path)
        if stream_df is not None:
            return {
                "path": path,
                "kind": "table",
                "df": self.normalize_pickle_dataframe(stream_df),
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
            }

        obj = pd.read_pickle(path)
        if isinstance(obj, pd.DataFrame):
            df = self.normalize_pickle_dataframe(obj)
            return {
                "path": path,
                "kind": "table",
                "df": df,
                "preview_only": False,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
            }
        if isinstance(obj, pd.Series):
            df = self.normalize_pickle_dataframe(obj.to_frame(name=obj.name or "value"))
            return {
                "path": path,
                "kind": "table",
                "df": df,
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
            }
        if isinstance(obj, np.ndarray):
            if obj.ndim == 1:
                df = pd.DataFrame({"value": obj.tolist()})
            elif obj.ndim == 2:
                df = pd.DataFrame(obj)
            else:
                text = f"NumPy array shape={obj.shape}, dtype={obj.dtype}\n\n"
                text += np.array2string(obj, threshold=200, edgeitems=3)
                return {"path": path, "kind": "text", "text": text}
            return {
                "path": path,
                "kind": "table",
                "df": self.normalize_pickle_dataframe(df),
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
            }
        if isinstance(obj, list) and obj and all(isinstance(item, dict) for item in obj):
            return {
                "path": path,
                "kind": "table",
                "df": self.normalize_pickle_dataframe(pd.DataFrame(obj)),
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
            }
        if isinstance(obj, dict):
            try:
                df = pd.DataFrame.from_dict(obj)
                if not df.empty:
                    return {
                        "path": path,
                        "kind": "table",
                        "df": self.normalize_pickle_dataframe(df),
                        "preview_only": True,
                        "excel_sheets": [],
                        "excel_sheet_index": 0,
                        "excel_sheet": 0,
                    }
            except Exception:
                pass
            # Fallback for scalar/mixed dicts: show as key/value rows.
            kv_rows = [{"key": str(k), "value": pprint.pformat(v, width=80, compact=True)} for k, v in obj.items()]
            return {
                "path": path,
                "kind": "table",
                "df": pd.DataFrame(kv_rows),
                "preview_only": True,
                "excel_sheets": [],
                "excel_sheet_index": 0,
                "excel_sheet": 0,
            }
        text = pprint.pformat(obj, width=100, compact=False)
        return {"path": path, "kind": "text", "text": text}

    def is_pickle_loading_allowed(self) -> bool:
        if self.allow_unsafe_pickle_session:
            return True
        value = os.environ.get(ALLOW_UNSAFE_PICKLE_ENV, "")
        return value.strip().lower() in {"1", "true", "yes", "on"}

    def read_pickle_stream_preview(self, path: Path) -> pd.DataFrame | None:
        """Read pickle files written as repeated pickle.dump(tuple_row, file)."""
        rows: list[Any] = []
        with path.open("rb") as fh:
            for _ in range(MAX_RENDER_ROWS + 1):
                try:
                    rows.append(pickle.load(fh))
                except EOFError:
                    break
                except Exception:
                    return None

        if len(rows) < 2:
            return None
        if not all(isinstance(r, tuple) for r in rows):
            return None

        header = rows[0]
        data_rows = rows[1:]
        if not all(isinstance(v, str) for v in header):
            return None
        width = len(header)
        if width == 0:
            return None

        # Keep only rows that match header width to avoid malformed records.
        filtered = [r for r in data_rows if isinstance(r, tuple) and len(r) == width]
        if not filtered:
            return None

        col_names = [h if h and str(h).strip() else f"col_{i+1}" for i, h in enumerate(header)]
        return pd.DataFrame(filtered, columns=col_names)

    def normalize_pickle_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        # Preserve meaningful index content so pickle tables don't look header-only.
        if df.empty:
            return df
        try:
            if not isinstance(df.index, pd.RangeIndex):
                return df.reset_index()
            if df.shape[1] == 0:
                return df.reset_index()
        except Exception:
            return df
        return df

    def prepare_excel_text_result(self, path: Path) -> dict[str, Any]:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_names = list(workbook.sheetnames)
            sheet_name = sheet_names[0] if sheet_names else ""
            if not sheet_name:
                return {"path": path, "kind": "text", "text": "[Empty workbook]"}
            ws = workbook[sheet_name]
            rows = ws.iter_rows(
                min_row=1,
                max_row=MAX_EXCEL_RENDER_ROWS,
                min_col=1,
                max_col=MAX_EXCEL_RENDER_COLS,
                values_only=True,
            )
            lines: list[str] = []
            for row in rows:
                cells = []
                for value in row:
                    if value is None:
                        cells.append("")
                    else:
                        text = str(value).replace("\r", " ").replace("\n", " ")
                        if len(text) > MAX_CELL_DISPLAY_LEN:
                            text = text[: MAX_CELL_DISPLAY_LEN - 3] + "..."
                        cells.append(text)
                line = " | ".join(cells).rstrip(" |")
                if len(line) > MAX_TEXT_PREVIEW_LINE_LEN:
                    line = line[: MAX_TEXT_PREVIEW_LINE_LEN - 3] + "..."
                lines.append(line)
            preview = "\n".join(lines).strip()
            if not preview:
                preview = "[No visible values in preview window]"
            header = f"Workbook: {path.name}\nSheet: {sheet_name}\n\n"
            df = self.read_excel_sheet(path, sheet_name)
            return {
                "path": path,
                "kind": "text",
                "text": header + preview,
                "df": df,
                "excel_sheets": sheet_names,
                "excel_sheet": sheet_name,
                "excel_sheet_index": 0,
            }
        finally:
            workbook.close()

    def prepare_image_result(self, path: Path) -> dict[str, Any]:
        with Image.open(path) as img:
            frame_count = getattr(img, "n_frames", 1)
            if frame_count > 1:
                img.seek(0)
            rgb = ImageOps.exif_transpose(img.convert("RGB"))
            width, height = rgb.size
            term_width = 100
            # Two pixels vertically per rendered char row.
            scale = min(1.0, term_width / max(1, width))
            new_w = max(2, int(width * scale))
            new_h = max(2, int(height * scale))
            rgb = rgb.resize((new_w, new_h), Image.Resampling.BILINEAR)
            pixels = np.array(rgb)

        ansi_lines: list[str] = []
        for y in range(0, pixels.shape[0] - 1, 2):
            parts: list[str] = []
            for x in range(pixels.shape[1]):
                top = pixels[y, x]
                bottom = pixels[y + 1, x]
                parts.append(
                    f"\x1b[38;2;{top[0]};{top[1]};{top[2]}m"
                    f"\x1b[48;2;{bottom[0]};{bottom[1]};{bottom[2]}m▀"
                )
            parts.append("\x1b[0m")
            ansi_lines.append("".join(parts))

        meta = (
            f"Image: {path.name}\n"
            f"Original: {width}x{height}\n"
            f"Preview: {new_w}x{new_h}\n"
            f"Mode: RGB\n"
            f"Frames: {frame_count}\n"
        )
        return {
            "path": path,
            "kind": "image",
            "ansi": "\n".join(ansi_lines),
            "meta": meta,
        }

    def apply_loaded_result(self, result: dict[str, Any]) -> None:
        path = result["path"]
        kind = result["kind"]
        self.current_path = path
        self.current_kind = kind
        self.current_df = None

        if kind == "table":
            self.lazy_delimited_context = result.get("lazy_delimited")
            self.visible_columns = None
            self.current_sort = None
            if path.suffix.lower() in {".xlsx", ".xls"}:
                df = result["df"]
                preview_text = df.to_string(index=False, max_rows=MAX_RENDER_ROWS, max_cols=MAX_RENDER_COLS)
                self.current_kind = "text"
                self.current_table_preview_only = False
                self.show_text(preview_text)
                self.query_one("#status", Static).update(
                    f"Loaded Excel preview as text: {path} | Sheet: {result.get('excel_sheet', 0)}"
                )
                self.post_message(FileLoaded(path))
                return
            self.current_table_preview_only = False
            self.current_excel_sheets = result.get("excel_sheets", [])
            self.current_excel_sheet_index = int(result.get("excel_sheet_index", 0))
            self.current_excel_sheet = result.get("excel_sheet", 0)
            try:
                self.render_table(result["df"])
            except Exception as exc:
                self.handle_load_error(path, exc)
                return
            if path.suffix.lower() in SQLITE_EXTENSIONS:
                self.show_db_schema_tree(result.get("db_schema"), engine="sqlite")
            elif path.suffix.lower() in DUCKDB_EXTENSIONS:
                self.show_db_schema_tree(result.get("db_schema"), engine="duckdb")
            elif path.suffix.lower() in FIREBIRD_EXTENSIONS:
                self.show_db_schema_tree(result.get("db_schema"), engine="firebird")
            if self.current_excel_sheets and len(self.current_excel_sheets) > 1:
                self.notify(
                    f"Excel file has {len(self.current_excel_sheets)} sheets. "
                    f"Loaded first sheet: {self.current_excel_sheet}"
                )
        elif kind == "log":
            self.current_table_preview_only = False
            self.current_excel_sheets = []
            self.current_excel_sheet_index = 0
            self.current_excel_sheet = 0
            self.current_log_text = result.get("text", "")
            self.log_text_colored_ready = False
            self.show_colored_log_text(self.current_log_text)
            if result.get("truncated"):
                self.notify("Log preview truncated for performance.")
        elif kind == "image":
            self.show_image(result.get("ansi", ""), result.get("meta", ""))
            self.current_excel_sheets = []
            self.current_excel_sheet_index = 0
        else:
            if path.suffix.lower() in {".xlsx", ".xls"}:
                self.current_table_preview_only = False
                self.current_excel_sheets = result.get("excel_sheets", [])
                self.current_excel_sheet_index = int(result.get("excel_sheet_index", 0))
                self.current_excel_sheet = result.get("excel_sheet", 0)
                df = result.get("df")
                if isinstance(df, pd.DataFrame):
                    try:
                        self.render_table(df)
                    except Exception:
                        pass
            else:
                self.current_excel_sheets = []
                self.current_excel_sheet_index = 0
            text = result.get("text", "")
            if not text.strip():
                text = "[Empty preview]"
            self.show_text(text)
            if kind == "json":
                self.show_json_tree(result.get("json_obj"))
            elif path.suffix.lower() in XML_EXTENSIONS:
                self.show_json_tree(result.get("json_obj"))
            if path.suffix.lower() in MARKDOWN_EXTENSIONS:
                self.show_markdown(text)
            elif path.suffix.lower() in XML_EXTENSIONS:
                self.show_markdown(self.xml_to_markdown_preview(path, text))

        self.post_message(FileLoaded(path))
        self.update_status_details()

    def handle_load_error(self, path: Path, exc: Exception) -> None:
        self.notify(f"Load failed: {exc}", severity="error")
        self.query_one("#status", Static).update(f"Failed to load: {path}")
        self.show_text(f"Failed to load file:\n{path}\n\n{type(exc).__name__}: {exc}")

    def render_table(self, df: pd.DataFrame) -> None:
        if self.lazy_delimited_context is not None:
            self.current_df_source = None
            self.current_df = df
        else:
            self.current_df_source = df
            self.current_df = df
        self.update_table_text_summary(df)
        self.table_page = 0
        self.apply_table_filter()
        self.render_current_table_page()
        self.render_schema_profile()
        self.update_status_details()

    def apply_table_filter(self) -> None:
        if self.lazy_delimited_context is not None:
            query = self.table_filter_query.strip()
            if not query:
                self.lazy_delimited_context["filtered_indices"] = None
                self.lazy_delimited_context["filtered_total_rows"] = None
                self.load_lazy_delimited_page(0)
                return
            self.apply_lazy_delimited_filter(query)
            self.render_schema_profile()
            return
        if self.current_df_source is None:
            self.current_df = None
            return
        query = self.table_filter_query.strip()
        if not query:
            self.current_df = self.current_df_source
            return
        filtered = self.filter_dataframe_with_query(self.current_df_source, query)
        self.current_df = filtered
        self.apply_sort_if_needed()
        self.render_schema_profile()

    def filter_dataframe_with_query(self, df: pd.DataFrame, query: str, notify_errors: bool = True) -> pd.DataFrame:
        # Column-aware filters:
        # col=value  -> contains (case-insensitive)
        # col==value -> exact (case-insensitive)
        # col!=value -> not exact (case-insensitive)
        # col~regex  -> regex match
        for op in ("!=", "==", "~", "="):
            if op in query:
                left, right = query.split(op, 1)
                col_name = left.strip()
                raw_value = right.strip()
                if not col_name:
                    break
                col = self.resolve_column_name(df, col_name)
                if col is None:
                    if notify_errors:
                        self.notify(f"Unknown column: {col_name}")
                    return df.iloc[0:0]
                series = df[col].astype(str)
                if op == "=":
                    v = raw_value.lower()
                    mask = series.str.lower().str.contains(re.escape(v), na=False)
                    return df[mask]
                if op == "==":
                    v = raw_value.lower()
                    mask = series.str.lower() == v
                    return df[mask]
                if op == "!=":
                    v = raw_value.lower()
                    mask = series.str.lower() != v
                    return df[mask]
                if op == "~":
                    try:
                        mask = series.str.contains(raw_value, regex=True, na=False)
                        return df[mask]
                    except re.error as exc:
                        if notify_errors:
                            self.notify(f"Invalid regex: {exc}")
                        return df.iloc[0:0]
        squery = query.lower()
        mask = df.apply(lambda row: any(squery in str(v).lower() for v in row.values), axis=1)
        return df[mask]

    def resolve_column_name(self, df: pd.DataFrame, name: str) -> str | None:
        for col in df.columns:
            if str(col) == name:
                return str(col)
        lname = name.lower()
        for col in df.columns:
            if str(col).lower() == lname:
                return str(col)
        return None

    def apply_sort_if_needed(self) -> None:
        if self.current_sort is None or self.current_df is None:
            return
        col, ascending = self.current_sort
        if col in self.current_df.columns:
            self.current_df = self.sort_dataframe_by_column(self.current_df, col, ascending)

    def sort_dataframe_by_column(self, df: pd.DataFrame, col: str, ascending: bool) -> pd.DataFrame:
        series = df[col]
        numeric = pd.to_numeric(series, errors="coerce")
        if numeric.notna().sum() > 0:
            key_series = numeric
        else:
            dt = pd.to_datetime(series, errors="coerce")
            if dt.notna().sum() > 0:
                key_series = dt
            else:
                key_series = series.astype(str).str.lower()
        temp = df.copy()
        temp["__sort_key__"] = key_series
        temp = temp.sort_values(by="__sort_key__", ascending=ascending, kind="mergesort", na_position="last")
        return temp.drop(columns=["__sort_key__"])

    def update_table_text_summary(self, df: pd.DataFrame) -> None:
        path = self.current_path
        size_text = "unknown"
        if path is not None and path.exists():
            size_text = f"{path.stat().st_size:,} bytes"
        columns = [str(c) for c in df.columns]
        lines = [
            "Table Summary",
            f"File: {path}" if path is not None else "File: [unknown]",
            f"Size: {size_text}",
            f"Rows: {len(df):,}",
            f"Columns: {len(columns):,}",
            "",
            "Column Names:",
        ]
        if columns:
            lines.extend([f"- {name}" for name in columns])
        else:
            lines.append("- [none]")
        text = self.query_one(TextArea)
        text.text = "\n".join(lines)

    def render_current_table_page(self) -> None:
        if self.current_df is None:
            return
        df = self.current_df
        self.query_one(TabbedContent).active = "table_tab"
        table = self.query_one(DataTable)
        table.clear(columns=True)
        table.cursor_type = "cell"

        if self.lazy_delimited_context is not None:
            render_df = df
        else:
            start = self.table_page * self.table_page_size
            end = start + self.table_page_size
            render_df = df.iloc[start:end]
        render_cols = self.get_render_columns(df)
        render_df = render_df.loc[:, render_cols] if render_cols else render_df.iloc[:, :0]
        if render_df.shape[1] == 0:
            table.add_column("value", key="col_1")
            self.update_pager()
            return

        columns = [str(c) for c in render_df.columns]
        # DataTable column keys must be unique even when headers repeat.
        for idx, label in enumerate(columns, start=1):
            safe_label = label if label.strip() else f"col_{idx}"
            if self.current_sort is not None and self.current_sort[0] == label:
                direction = "▲" if self.current_sort[1] else "▼"
                safe_label = f"{safe_label} {direction}"
            table.add_column(safe_label, key=f"col_{idx}")
        for row in render_df.itertuples(index=False, name=None):
            table.add_row(*[self.format_cell_for_display(v) for v in row])
        self.update_pager()

    @on(Input.Submitted, "#table_filter_input")
    def on_table_filter_submitted(self, event: Input.Submitted) -> None:
        self.table_filter_query = event.value or ""
        if self.lazy_delimited_context is None and self.current_df_source is None:
            return
        self.table_page = 0
        self.apply_table_filter()
        self.render_current_table_page()
        self.update_status_details()

    @on(Button.Pressed, "#table_filter_clear")
    def on_table_filter_clear(self) -> None:
        self.query_one("#table_filter_input", Input).value = ""
        self.table_filter_query = ""
        if self.lazy_delimited_context is None and self.current_df_source is None:
            return
        self.table_page = 0
        self.apply_table_filter()
        self.render_current_table_page()
        self.update_status_details()


    def update_pager(self) -> None:
        if self.current_df is None:
            self.query_one("#pager", Horizontal).display = False
            return
        total_rows = (
            int(
                self.lazy_delimited_context.get("filtered_total_rows")
                if self.lazy_delimited_context is not None and self.lazy_delimited_context.get("filtered_total_rows") is not None
                else self.lazy_delimited_context["total_rows"]
            )
            if self.lazy_delimited_context is not None
            else len(self.current_df)
        )
        total_pages = max(1, (total_rows + self.table_page_size - 1) // self.table_page_size)
        self.table_page = max(0, min(self.table_page, total_pages - 1))
        pager = self.query_one("#pager", Horizontal)
        pager.display = True
        self.query_one("#page_info", Static).update(
            f"Page {self.table_page + 1}/{total_pages} | Rows: {total_rows} | Page size: {self.table_page_size}"
        )
        self.query_one("#page_first", Button).disabled = self.table_page <= 0
        self.query_one("#page_prev", Button).disabled = self.table_page <= 0
        self.query_one("#page_next", Button).disabled = self.table_page >= total_pages - 1
        self.query_one("#page_last", Button).disabled = self.table_page >= total_pages - 1

    @on(Button.Pressed, "#page_first")
    def page_first(self) -> None:
        if self.current_df is None:
            return
        if self.table_page != 0:
            self.table_page = 0
            if self.lazy_delimited_context is not None:
                self.load_lazy_page_async(self.table_page)
            else:
                self.render_current_table_page()

    @on(Button.Pressed, "#page_prev")
    def page_prev(self) -> None:
        if self.current_df is None:
            return
        if self.table_page > 0:
            self.table_page -= 1
            if self.lazy_delimited_context is not None:
                self.load_lazy_page_async(self.table_page)
            else:
                self.render_current_table_page()

    @on(Button.Pressed, "#page_next")
    def page_next(self) -> None:
        if self.current_df is None:
            return
        total_rows = (
            int(
                self.lazy_delimited_context.get("filtered_total_rows")
                if self.lazy_delimited_context is not None and self.lazy_delimited_context.get("filtered_total_rows") is not None
                else self.lazy_delimited_context["total_rows"]
            )
            if self.lazy_delimited_context is not None
            else len(self.current_df)
        )
        total_pages = max(1, (total_rows + self.table_page_size - 1) // self.table_page_size)
        if self.table_page < total_pages - 1:
            self.table_page += 1
            if self.lazy_delimited_context is not None:
                self.load_lazy_page_async(self.table_page)
            else:
                self.render_current_table_page()

    @on(Button.Pressed, "#page_last")
    def page_last(self) -> None:
        if self.current_df is None:
            return
        total_rows = (
            int(
                self.lazy_delimited_context.get("filtered_total_rows")
                if self.lazy_delimited_context is not None and self.lazy_delimited_context.get("filtered_total_rows") is not None
                else self.lazy_delimited_context["total_rows"]
            )
            if self.lazy_delimited_context is not None
            else len(self.current_df)
        )
        total_pages = max(1, (total_rows + self.table_page_size - 1) // self.table_page_size)
        last_page = total_pages - 1
        if self.table_page != last_page:
            self.table_page = last_page
            if self.lazy_delimited_context is not None:
                self.load_lazy_page_async(self.table_page)
            else:
                self.render_current_table_page()

    def format_cell_for_display(self, value: Any) -> str:
        if pd.isna(value):
            return ""
        text = str(value).replace("\r", " ").replace("\n", " ")
        if len(text) > MAX_CELL_DISPLAY_LEN:
            return text[: MAX_CELL_DISPLAY_LEN - 3] + "..."
        return text

    def read_xml_with_fallback(self, path: Path) -> str:
        raw_text = self.read_text_with_fallback(path)
        try:
            root = ET.fromstring(raw_text)
            normalized = ET.tostring(root, encoding="unicode")
            return minidom.parseString(normalized.encode("utf-8")).toprettyxml(indent="  ")
        except Exception:
            return raw_text

    def read_html_with_fallback(self, path: Path) -> str:
        raw_text = self.read_text_with_fallback(path)
        try:
            soup = BeautifulSoup(raw_text, "html.parser")
            for tag in soup(["script", "style", "noscript"]):
                tag.decompose()

            root = soup.body if soup.body is not None else soup
            lines: list[str] = []

            title = soup.title.get_text(" ", strip=True) if soup.title else ""
            if title:
                lines.append(f"Title: {title}")
                lines.append("")

            for node in root.find_all(
                ["h1", "h2", "h3", "h4", "h5", "h6", "p", "li", "a", "tr", "pre", "code"]
            ):
                text = node.get_text(" ", strip=True)
                if not text:
                    continue
                name = node.name or ""
                if name.startswith("h") and len(name) == 2 and name[1].isdigit():
                    level = int(name[1])
                    lines.append(f"{'#' * level} {text}")
                    lines.append("")
                elif name == "li":
                    lines.append(f"- {text}")
                elif name == "a":
                    href = node.get("href", "")
                    lines.append(f"{text} ({href})" if href else text)
                elif name == "tr":
                    cells = [c.get_text(" ", strip=True) for c in node.find_all(["th", "td"])]
                    if any(cells):
                        lines.append(" | ".join(cells))
                elif name in {"pre", "code"}:
                    lines.append(text)
                else:
                    lines.append(text)
                    lines.append("")

            rendered = "\n".join(lines).strip()
            if rendered:
                return rendered
        except Exception:
            pass

        # Fallback for malformed HTML.
        return re.sub(r">\s*<", ">\n<", raw_text)

    def read_pdf_preview(self, path: Path) -> str:
        reader = PdfReader(str(path))
        total_pages = len(reader.pages)
        page_limit = min(total_pages, MAX_PDF_PAGES)
        parts: list[str] = [f"PDF: {path.name}", f"Pages: {total_pages}", ""]
        char_count = 0
        for page_idx in range(page_limit):
            page = reader.pages[page_idx]
            text = page.extract_text() or ""
            if not text.strip():
                text = "[No extractable text on this page]"
            chunk = f"--- Page {page_idx + 1} ---\n{text}\n"
            if char_count + len(chunk) > MAX_PDF_CHARS:
                remaining = MAX_PDF_CHARS - char_count
                if remaining > 0:
                    parts.append(chunk[:remaining])
                parts.append("\n[Preview truncated]")
                break
            parts.append(chunk)
            char_count += len(chunk)
        if total_pages > page_limit:
            parts.append(f"\n[Showing first {page_limit} pages]")
        return "\n".join(parts).strip()

    def load_json(self, path: Path) -> None:
        raw_text = self.read_text_with_fallback(path)
        data = json.loads(raw_text)
        self.show_text(json.dumps(data, indent=2, ensure_ascii=False))

    def load_text(self, path: Path) -> None:
        self.show_text(self.read_text_with_fallback(path))

    def read_text_with_fallback(self, path: Path) -> str:
        encodings = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
        last_error: UnicodeDecodeError | None = None
        for encoding in encodings:
            try:
                return path.read_text(encoding=encoding)
            except UnicodeDecodeError as exc:
                last_error = exc
        if last_error is not None:
            raise ValueError(f"Could not decode file as text with {encodings}") from last_error
        raise ValueError("Could not decode file as text.")

    def read_delimited_with_fallback(self, path: Path, sep: str) -> pd.DataFrame:
        encodings = ("utf-8", "utf-8-sig", "cp1252", "latin-1")
        last_error: UnicodeDecodeError | None = None
        for encoding in encodings:
            try:
                return pd.read_csv(path, sep=sep, encoding=encoding, nrows=MAX_TABULAR_PREVIEW_ROWS)
            except UnicodeDecodeError as exc:
                last_error = exc
            except pd.errors.ParserError:
                continue

        for encoding in encodings:
            try:
                return pd.read_csv(path, sep=None, engine="python", encoding=encoding, nrows=MAX_TABULAR_PREVIEW_ROWS)
            except UnicodeDecodeError as exc:
                last_error = exc
            except pd.errors.ParserError:
                continue

        for encoding in encodings:
            try:
                return pd.read_csv(
                    path,
                    sep=None,
                    engine="python",
                    encoding=encoding,
                    on_bad_lines="skip",
                    nrows=MAX_TABULAR_PREVIEW_ROWS,
                )
            except UnicodeDecodeError as exc:
                last_error = exc
            except pd.errors.ParserError:
                continue
        if last_error is not None:
            raise ValueError(f"Could not decode delimited file with {encodings}") from last_error
        raise ValueError("Could not parse delimited file.")

    def read_gzip_text_with_fallback(self, path: Path) -> str:
        encodings = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
        last_error: UnicodeDecodeError | None = None
        for encoding in encodings:
            try:
                with gzip.open(path, mode="rt", encoding=encoding) as fh:
                    return fh.read()
            except UnicodeDecodeError as exc:
                last_error = exc
        if last_error is not None:
            raise ValueError(f"Could not decode gzip text with {encodings}") from last_error
        raise ValueError("Could not decode gzip text.")

    def read_gzip_delimited_with_fallback(self, path: Path, sep: str) -> pd.DataFrame:
        encodings = ("utf-8", "utf-8-sig", "cp1252", "latin-1")
        last_error: UnicodeDecodeError | None = None
        for encoding in encodings:
            try:
                return pd.read_csv(
                    path,
                    sep=sep,
                    encoding=encoding,
                    compression="gzip",
                    nrows=MAX_TABULAR_PREVIEW_ROWS,
                )
            except UnicodeDecodeError as exc:
                last_error = exc
            except pd.errors.ParserError:
                continue

        for encoding in encodings:
            try:
                return pd.read_csv(
                    path,
                    sep=None,
                    engine="python",
                    encoding=encoding,
                    compression="gzip",
                    nrows=MAX_TABULAR_PREVIEW_ROWS,
                )
            except UnicodeDecodeError as exc:
                last_error = exc
            except pd.errors.ParserError:
                continue

        for encoding in encodings:
            try:
                return pd.read_csv(
                    path,
                    sep=None,
                    engine="python",
                    encoding=encoding,
                    compression="gzip",
                    on_bad_lines="skip",
                    nrows=MAX_TABULAR_PREVIEW_ROWS,
                )
            except UnicodeDecodeError as exc:
                last_error = exc
            except pd.errors.ParserError:
                continue
        if last_error is not None:
            raise ValueError(f"Could not decode gzip delimited file with {encodings}") from last_error
        raise ValueError("Could not parse gzip delimited file.")

    def read_ndjson_preview(self, content: str) -> list[Any]:
        records: list[Any] = []
        for line_no, line in enumerate(content.splitlines(), start=1):
            stripped = line.strip()
            if not stripped:
                continue
            try:
                records.append(json.loads(stripped))
            except json.JSONDecodeError as exc:
                records.append({"_parse_error": f"line {line_no}: {exc.msg}", "_raw": stripped})
            if len(records) >= MAX_RENDER_ROWS:
                break
        return records

    def read_arrow_preview(self, path: Path, format_name: str) -> pd.DataFrame:
        dataset = ds.dataset(str(path), format=format_name)
        rows_remaining = MAX_RENDER_ROWS
        batches: list[pa.RecordBatch] = []
        preview_columns = dataset.schema.names[:MAX_RENDER_COLS]
        for batch in dataset.to_batches(columns=preview_columns, batch_size=min(1000, MAX_RENDER_ROWS)):
            if rows_remaining <= 0:
                break
            if batch.num_rows > rows_remaining:
                batch = batch.slice(0, rows_remaining)
            batches.append(batch)
            rows_remaining -= batch.num_rows

        if not batches:
            return pd.DataFrame()

        table = pa.Table.from_batches(batches)
        return table.to_pandas()

    def read_avro_preview(self, path: Path) -> pd.DataFrame:
        rows: list[dict[str, Any]] = []
        with path.open("rb") as fh:
            for record in avro_reader(fh):
                rows.append({k: v for idx, (k, v) in enumerate(record.items()) if idx < MAX_RENDER_COLS})
                if len(rows) >= MAX_RENDER_ROWS:
                    break
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame(rows)

    def read_excel_sheet(self, path: Path, sheet_name: str | int) -> pd.DataFrame:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = workbook[sheet_name] if isinstance(sheet_name, str) else workbook.worksheets[int(sheet_name)]
            # First pass: find where real data starts inside a generous scan area.
            first_data_row = None
            first_data_col = None
            scan_iter = ws.iter_rows(
                min_row=1,
                max_row=MAX_EXCEL_SCAN_ROWS,
                min_col=1,
                max_col=MAX_EXCEL_SCAN_COLS,
                values_only=True,
            )
            for r_idx, row in enumerate(scan_iter, start=1):
                for c_idx, cell in enumerate(row, start=1):
                    if cell is not None and str(cell).strip() != "":
                        first_data_row = r_idx
                        first_data_col = c_idx
                        break
                if first_data_row is not None:
                    break

            if first_data_row is None or first_data_col is None:
                return pd.DataFrame()

            # Second pass: read a bounded window from detected start.
            rows_iter = ws.iter_rows(
                min_row=first_data_row,
                max_row=first_data_row + MAX_EXCEL_RENDER_ROWS,
                min_col=first_data_col,
                max_col=first_data_col + MAX_EXCEL_RENDER_COLS - 1,
                values_only=True,
            )
            rows = [tuple(row) for row in rows_iter]
        finally:
            workbook.close()

        if not rows:
            return pd.DataFrame()

        header_row = rows[0]
        header = [f"col_{idx+1}" if v is None or str(v).strip() == "" else str(v) for idx, v in enumerate(header_row)]
        data = rows[1 : 1 + MAX_EXCEL_RENDER_ROWS]
        return pd.DataFrame(data, columns=header)

    def action_prev_sheet(self) -> None:
        self.switch_excel_sheet(-1)

    def action_next_sheet(self) -> None:
        self.switch_excel_sheet(1)

    def switch_excel_sheet(self, delta: int) -> None:
        if self.current_path is None or self.current_path.suffix.lower() not in {".xlsx", ".xls"}:
            self.notify("Sheet switching is available only for Excel files.")
            return
        if not self.current_excel_sheets:
            self.notify("No Excel sheets available.")
            return
        self.switch_excel_sheet_worker(delta)

    @work(thread=True)
    def switch_excel_sheet_worker(self, delta: int) -> None:
        assert self.current_path is not None
        next_index = (self.current_excel_sheet_index + delta) % len(self.current_excel_sheets)
        next_sheet = self.current_excel_sheets[next_index]
        try:
            df = self.read_excel_sheet(self.current_path, next_sheet)
            self.call_from_thread(self.apply_switched_sheet, next_index, next_sheet, df)
        except Exception as exc:
            self.call_from_thread(self.notify, f"Sheet switch failed: {exc}", severity="error")

    def apply_switched_sheet(self, index: int, sheet_name: str, df: pd.DataFrame) -> None:
        self.current_excel_sheet_index = index
        self.current_excel_sheet = sheet_name
        self.current_table_preview_only = False
        self.render_table(df)
        self.query_one("#status", Static).update(f"Loaded: {self.current_path} | Sheet: {self.current_excel_sheet}")
        self.notify(f"Switched to sheet: {self.current_excel_sheet}")

    def show_text(self, content: str) -> None:
        self.query_one(TabbedContent).active = "text_tab"
        log_preview = self.query_one("#log_text_preview", Static)
        text = self.query_one(TextArea)
        text.display = True
        log_preview.display = False
        text.soft_wrap = False
        is_makefile = False
        if self.current_path is not None:
            suffix = self.current_path.suffix.lower()
            is_makefile = self.current_path.name.lower() in MAKEFILE_NAMES or suffix == ".mk"
            if suffix in {".py", ".pyw"}:
                text.language = "python"
            elif suffix == ".sql":
                text.language = "sql"
            elif is_makefile:
                text.language = None
                text.text = content
                text.display = False
                log_preview.display = True
                log_preview.update(self.colorize_makefile_text(content))
                return
            elif suffix in MARKDOWN_EXTENSIONS:
                text.language = "markdown"
            elif suffix in YAML_EXTENSIONS:
                text.language = "yaml"
            elif suffix in INI_EXTENSIONS:
                text.language = None
                text.text = content
                text.display = False
                log_preview.display = True
                log_preview.update(self.colorize_ini_text(content))
                return
            elif suffix in TOML_EXTENSIONS:
                text.language = None
                text.text = content
                text.display = False
                log_preview.display = True
                log_preview.update(self.colorize_toml_text(content))
                return
            elif suffix in ENV_EXTENSIONS:
                text.language = None
                text.text = content
                text.display = False
                log_preview.display = True
                log_preview.update(self.colorize_env_text(content))
                return
            elif suffix in NDJSON_EXTENSIONS:
                text.language = "json"
            elif suffix == ".sh":
                text.language = "bash"
            elif suffix == ".bat":
                text.language = None
                text.text = content
                text.display = False
                log_preview.display = True
                log_preview.update(self.colorize_bat_text(content))
                return
            else:
                text.language = None
        else:
            text.language = None
        text.text = content
        if self.text_search_highlight_all and self.last_text_search:
            self.text_search_matches = self.compute_text_matches(content, self.last_text_search)
            if self.text_search_matches:
                if self.text_search_index < 0:
                    self.text_search_index = 0
                else:
                    self.text_search_index = min(self.text_search_index, len(self.text_search_matches) - 1)
                self.render_search_highlight_preview()

    def show_colored_log_text(self, content: str) -> None:
        text = self.query_one(TextArea)
        log_preview = self.query_one("#log_text_preview", Static)
        text.display = False
        text.text = content
        log_preview.display = True
        log_preview.update(self.colorize_log_text(content))
        self.log_text_colored_ready = True

    @on(TabbedContent.TabActivated)
    def on_tab_activated(self, event: TabbedContent.TabActivated) -> None:
        active_tab = self.query_one(TabbedContent).active
        if active_tab != "table_tab":
            self.query_one("#pager", Horizontal).display = False
        elif self.current_df is not None:
            self.update_pager()
        if active_tab == "text_tab" and self.current_kind == "log":
            if self.current_log_text and not self.log_text_colored_ready:
                self.show_colored_log_text(self.current_log_text)
        self.update_context_shortcuts()

    def update_context_shortcuts(self) -> None:
        active_tab = self.query_one(TabbedContent).active
        hints = "Context: "
        if active_tab == "table_tab":
            hints += "g Jump Page | [ / ] Page Size | e Edit Cell | Enter Apply Filter | Alt+S/F7 Sort | Ctrl+K Cols | Alt+P/F6 Full"
        elif active_tab == "text_tab":
            hints += "/ Search | F3 Next | Shift+F3 Prev | Ctrl+H Highlight"
            if self.current_kind == "log":
                hints += " | Timestamp Highlight"
        elif active_tab == "json_tree_tab":
            hints += "Browse JSON structure"
        elif active_tab == "markdown_tab":
            hints += "Rendered markdown preview"
        elif active_tab == "image_tab":
            hints += "Image preview"
        else:
            hints += "Open a file from the tree"
        self.query_one("#context_shortcuts", Static).update(hints)

    def colorize_log_text(self, content: str) -> Text:
        rendered = Text()
        ts_pattern = re.compile(r"^(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}(?:[.,]\d+)?)")
        for line in content.splitlines():
            match = ts_pattern.match(line)
            if match:
                ts = match.group(1)
                rendered.append(ts, style="bold cyan")
                rendered.append(line[len(ts):])
            else:
                rendered.append(line)
            rendered.append("\n")
        return rendered

    def colorize_bat_text(self, content: str) -> Text:
        rendered = Text()
        keyword_pattern = re.compile(
            r"\b(@echo|echo|setlocal|endlocal|set|if|for|in|do|goto|call|shift|exit)\b",
            re.IGNORECASE,
        )
        var_pattern = re.compile(r"%[^%\s]+%|%%[A-Za-z]")
        rem_pattern = re.compile(r"^\s*(rem\b|::)", re.IGNORECASE)

        for line in content.splitlines():
            if rem_pattern.match(line):
                rendered.append(line, style="italic #6b7280")
                rendered.append("\n")
                continue

            cursor = 0
            for match in re.finditer(r"%[^%\s]+%|%%[A-Za-z]|\b(?:@echo|echo|setlocal|endlocal|set|if|for|in|do|goto|call|shift|exit)\b", line, re.IGNORECASE):
                start, end = match.span()
                if start > cursor:
                    rendered.append(line[cursor:start])
                token = match.group(0)
                if var_pattern.fullmatch(token):
                    rendered.append(token, style="bold #22c55e")
                elif keyword_pattern.fullmatch(token):
                    rendered.append(token, style="bold #60a5fa")
                else:
                    rendered.append(token)
                cursor = end
            if cursor < len(line):
                rendered.append(line[cursor:])
            rendered.append("\n")
        return rendered

    def colorize_ini_text(self, content: str) -> Text:
        rendered = Text()
        for line in content.splitlines():
            stripped = line.strip()
            if not stripped:
                rendered.append("\n")
                continue
            if stripped.startswith(("#", ";")):
                rendered.append(line, style="italic #6b7280")
                rendered.append("\n")
                continue
            if stripped.startswith("[") and stripped.endswith("]"):
                rendered.append(line, style="bold #60a5fa")
                rendered.append("\n")
                continue
            if "=" in line:
                key, value = line.split("=", 1)
                rendered.append(key, style="bold #22c55e")
                rendered.append("=")
                rendered.append(value, style="#e5e7eb")
                rendered.append("\n")
                continue
            rendered.append(line)
            rendered.append("\n")
        return rendered

    def colorize_toml_text(self, content: str) -> Text:
        rendered = Text()
        number_pattern = re.compile(r"^-?\d+(\.\d+)?$")
        for line in content.splitlines():
            stripped = line.strip()
            if not stripped:
                rendered.append("\n")
                continue
            if stripped.startswith("#"):
                rendered.append(line, style="italic #6b7280")
                rendered.append("\n")
                continue
            if stripped.startswith("[") and stripped.endswith("]"):
                rendered.append(line, style="bold #60a5fa")
                rendered.append("\n")
                continue
            if "=" in line:
                key, value = line.split("=", 1)
                v = value.strip()
                rendered.append(key, style="bold #22c55e")
                rendered.append(" = ")
                if v.lower() in {"true", "false"}:
                    rendered.append(v, style="bold #f59e0b")
                elif number_pattern.match(v):
                    rendered.append(v, style="#f59e0b")
                elif v.startswith('"') and v.endswith('"'):
                    rendered.append(v, style="#a78bfa")
                else:
                    rendered.append(value, style="#e5e7eb")
                rendered.append("\n")
                continue
            rendered.append(line)
            rendered.append("\n")
        return rendered

    def colorize_env_text(self, content: str) -> Text:
        rendered = Text()
        for line in content.splitlines():
            stripped = line.strip()
            if not stripped:
                rendered.append("\n")
                continue
            if stripped.startswith("#"):
                rendered.append(line, style="italic #6b7280")
                rendered.append("\n")
                continue
            if "=" in line:
                key, value = line.split("=", 1)
                rendered.append(key, style="bold #22c55e")
                rendered.append("=")
                v = value.strip()
                if v.lower() in {"true", "false"}:
                    rendered.append(value, style="bold #f59e0b")
                elif len(v) >= 2 and ((v[0] == '"' and v[-1] == '"') or (v[0] == "'" and v[-1] == "'")):
                    rendered.append(value, style="#a78bfa")
                else:
                    rendered.append(value, style="#e5e7eb")
                rendered.append("\n")
                continue
            rendered.append(line)
            rendered.append("\n")
        return rendered

    def colorize_makefile_text(self, content: str) -> Text:
        rendered = Text()
        target_pattern = re.compile(r"^\s*([A-Za-z0-9_.%/\\-]+)\s*:(?![=])")
        var_assign_pattern = re.compile(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*[:+?]?=")
        var_ref_pattern = re.compile(r"\$\(([^)]+)\)|\$\{([^}]+)\}")

        for line in content.splitlines():
            if not line.strip():
                rendered.append("\n")
                continue
            if line.lstrip().startswith("#"):
                rendered.append(line, style="italic #6b7280")
                rendered.append("\n")
                continue
            if line.startswith("\t"):
                # Recipe/command line
                rendered.append(line, style="#93c5fd")
                rendered.append("\n")
                continue

            m_target = target_pattern.match(line)
            m_var = var_assign_pattern.match(line)

            if m_target:
                start, end = m_target.span(1)
                rendered.append(line[:start])
                rendered.append(line[start:end], style="bold #f59e0b")
                rendered.append(line[end:])
            elif m_var:
                vstart, vend = m_var.span(1)
                rendered.append(line[:vstart])
                rendered.append(line[vstart:vend], style="bold #22c55e")
                rendered.append(line[vend:])
            else:
                rendered.append(line)

            # Overlay variable refs by rebuilding this line segment
            # for simple readability gains in common patterns.
            text_line = rendered.plain.splitlines()[-1] if rendered.plain else line
            if "$(" in text_line or "${" in text_line:
                # Remove just-appended line and re-add with var-ref styles.
                rendered = Text("\n".join(rendered.plain.splitlines()[:-1]) + ("\n" if rendered.plain.count("\n") else ""))
                cursor = 0
                for mv in var_ref_pattern.finditer(line):
                    s, e = mv.span()
                    if s > cursor:
                        rendered.append(line[cursor:s])
                    rendered.append(line[s:e], style="bold #a78bfa")
                    cursor = e
                if cursor < len(line):
                    rendered.append(line[cursor:])
            rendered.append("\n")
        return rendered

    def get_current_text_content(self) -> str:
        if self.current_kind == "log":
            return self.current_log_text
        return self.query_one(TextArea).text

    def compute_text_matches(self, content: str, query: str) -> list[int]:
        matches: list[int] = []
        low_content = content.lower()
        low_query = query.lower()
        if not low_query:
            return matches
        start = 0
        while True:
            pos = low_content.find(low_query, start)
            if pos < 0:
                break
            matches.append(pos)
            start = pos + max(1, len(low_query))
        return matches

    def offset_to_line_col(self, content: str, offset: int) -> tuple[int, int]:
        line = content.count("\n", 0, offset) + 1
        last_newline = content.rfind("\n", 0, offset)
        col = offset + 1 if last_newline < 0 else offset - last_newline
        return line, col

    def notify_current_search_match(self) -> None:
        if not self.text_search_matches or not self.last_text_search:
            return
        content = self.get_current_text_content()
        offset = self.text_search_matches[self.text_search_index]
        line, col = self.offset_to_line_col(content, offset)
        self.notify(f"Match {self.text_search_index + 1}/{len(self.text_search_matches)} at line {line}, col {col}")

    def render_search_highlight_preview(self) -> None:
        if not self.last_text_search:
            return
        content = self.get_current_text_content()
        if not content:
            return
        matches = self.text_search_matches or self.compute_text_matches(content, self.last_text_search)
        if not matches:
            return
        text = Text()
        cursor = 0
        qlen = len(self.last_text_search)
        active_offset = matches[self.text_search_index] if 0 <= self.text_search_index < len(matches) else -1
        for offset in matches:
            if offset > cursor:
                text.append(content[cursor:offset])
            token = content[offset : offset + qlen]
            style = "black on yellow"
            if offset == active_offset:
                style = "black on bright_magenta"
            text.append(token, style=style)
            cursor = offset + qlen
        if cursor < len(content):
            text.append(content[cursor:])
        preview = self.query_one("#log_text_preview", Static)
        editor = self.query_one(TextArea)
        editor.display = False
        preview.display = True
        preview.update(text)

    def restore_text_rendering(self) -> None:
        content = self.get_current_text_content()
        if self.current_kind == "log":
            self.show_colored_log_text(content)
            return
        self.show_text(content)

    def limit_log_preview(self, content: str) -> tuple[str, bool]:
        lines = content.splitlines()
        truncated = False
        if len(lines) > MAX_LOG_PREVIEW_LINES:
            lines = lines[:MAX_LOG_PREVIEW_LINES]
            truncated = True
        preview = "\n".join(lines)
        if len(preview) > MAX_LOG_PREVIEW_CHARS:
            preview = preview[:MAX_LOG_PREVIEW_CHARS]
            truncated = True
        if truncated:
            preview = f"{preview}\n\n[Preview truncated]"
        return preview, truncated

    def show_markdown(self, content: str) -> None:
        self.query_one("#markdown_preview", Static).update(Markdown(content))

    def show_json_tree(self, obj: Any) -> None:
        tree = self.query_one("#json_tree", Tree)
        tree.clear()
        tree.root.set_label("JSON")
        self.populate_json_tree(tree.root, obj)
        tree.root.expand()

    def show_db_schema_tree(self, schema_obj: Any, engine: str = "sqlite") -> None:
        tree = self.query_one("#db_schema_tree", Tree)
        tree.clear()
        tree.root.set_label("DB Schema")
        if not isinstance(schema_obj, dict):
            tree.root.add("No schema details")
            tree.root.expand()
            return
        tables = schema_obj.get("tables", {})
        views = schema_obj.get("views", [])
        indexes = schema_obj.get("indexes", [])
        triggers = schema_obj.get("triggers", [])

        if not isinstance(tables, dict):
            tables = {}
        if not isinstance(views, list):
            views = []
        if not isinstance(indexes, list):
            indexes = []
        if not isinstance(triggers, list):
            triggers = []

        if not tables and not views and not indexes and not triggers:
            tree.root.add("No schema details")
            tree.root.expand()
            return

        tables_node = tree.root.add("tables")
        if tables:
            for table_name, detail in tables.items():
                table_node = tables_node.add(
                    str(table_name),
                    data={"kind": "db_table", "name": str(table_name), "engine": engine},
                )

                cols_node = table_node.add("columns")
                columns = detail.get("columns", [])
                if columns:
                    for col in columns:
                        col_name = col.get("name", "?")
                        col_type = col.get("type", "")
                        nullable = "NOT NULL" if int(col.get("notnull", 0)) else "NULLABLE"
                        pk = " PK" if int(col.get("pk", 0)) else ""
                        default = col.get("dflt_value")
                        default_text = f" default={default}" if default is not None else ""
                        cols_node.add(f"{col_name} : {col_type} ({nullable}{pk}){default_text}")
                else:
                    cols_node.add("[none]")

                fk_node = table_node.add("foreign_keys")
                fks = detail.get("foreign_keys", [])
                if fks:
                    for fk in fks:
                        fk_node.add(
                            f'{fk.get("from")} -> {fk.get("table")}.{fk.get("to")} '
                            f'(on_update={fk.get("on_update")}, on_delete={fk.get("on_delete")})'
                        )
                else:
                    fk_node.add("[none]")

                idx_node = table_node.add("indexes")
                local_indexes = detail.get("indexes", [])
                if local_indexes:
                    for idx in local_indexes:
                        cols = ", ".join([str(c) for c in idx.get("columns", [])])
                        unique = "unique" if idx.get("unique") else "non-unique"
                        idx_node.add(f'{idx.get("name")} ({unique}) [{cols}]')
                else:
                    idx_node.add("[none]")
        else:
            tables_node.add("[none]")

        views_node = tree.root.add("views")
        if views:
            for view in views:
                vname = str(view.get("name", "?"))
                vnode = views_node.add(vname)
                sql = view.get("sql")
                if sql:
                    vnode.add(str(sql))
        else:
            views_node.add("[none]")

        indexes_node = tree.root.add("indexes")
        if indexes:
            for idx in indexes:
                cols = ", ".join([str(c) for c in idx.get("columns", [])])
                iname = str(idx.get("name", "?"))
                tname = str(idx.get("table", "?"))
                inode = indexes_node.add(f"{iname} on {tname} [{cols}]")
                sql = idx.get("sql")
                if sql:
                    inode.add(str(sql))
        else:
            indexes_node.add("[none]")

        triggers_node = tree.root.add("triggers")
        if triggers:
            for trg in triggers:
                tname = str(trg.get("name", "?"))
                table_name = str(trg.get("tbl_name", "?"))
                tnode = triggers_node.add(f"{tname} on {table_name}")
                sql = trg.get("sql")
                if sql:
                    tnode.add(str(sql))
        else:
            triggers_node.add("[none]")
        tree.root.expand()

    @on(Tree.NodeSelected, "#db_schema_tree")
    def on_db_schema_tree_selected(self, event: Tree.NodeSelected) -> None:
        data = event.node.data
        if not isinstance(data, dict):
            return
        if data.get("kind") != "db_table":
            return
        table_name = str(data.get("name", "")).strip()
        engine = str(data.get("engine", "")).strip().lower()
        if not table_name:
            return
        self.load_selected_db_table(table_name, engine)

    def load_selected_db_table(self, table_name: str, engine: str) -> None:
        if self.current_path is None:
            return
        try:
            sql = f'SELECT * FROM "{table_name}" LIMIT {MAX_RENDER_ROWS}'
            if engine == "sqlite":
                if self.current_path.suffix.lower() not in SQLITE_EXTENSIONS:
                    self.notify("Selected table belongs to a SQLite file.", severity="warning")
                    return
                conn = sqlite3.connect(str(self.current_path))
                try:
                    df = pd.read_sql_query(sql, conn)
                finally:
                    conn.close()
            elif engine == "duckdb":
                if self.current_path.suffix.lower() not in DUCKDB_EXTENSIONS:
                    self.notify("Selected table belongs to a DuckDB file.", severity="warning")
                    return
                try:
                    import duckdb  # type: ignore
                except Exception as exc:
                    raise RuntimeError("duckdb package is required to open .duckdb files") from exc
                conn = duckdb.connect(str(self.current_path), read_only=True)
                try:
                    df = conn.execute(sql).fetchdf()
                finally:
                    conn.close()
            elif engine == "firebird":
                if self.current_path.suffix.lower() not in FIREBIRD_EXTENSIONS:
                    self.notify("Selected table belongs to a Firebird file.", severity="warning")
                    return
                conn = self.connect_firebird(self.current_path)
                try:
                    fb_sql = f'SELECT * FROM "{table_name}" ROWS {MAX_RENDER_ROWS}'
                    df = self.firebird_query_df(conn, fb_sql)
                finally:
                    self.close_firebird(conn)
            else:
                self.notify("Unknown DB engine for selected table.", severity="warning")
                return
            self.current_kind = "table"
            self.lazy_delimited_context = None
            self.current_df_source = df
            self.current_df = df
            self.table_page = 0
            self.apply_table_filter()
            self.render_current_table_page()
            self.render_schema_profile()
            self.query_one(TabbedContent).active = "table_tab"
            self.update_status_details()
            self.notify(f"Loaded table: {table_name}")
        except Exception as exc:
            self.notify(f"Failed to load table {table_name}: {exc}", severity="error")

    def populate_json_tree(self, node, value: Any) -> None:
        if isinstance(value, dict):
            for key, item in value.items():
                if isinstance(item, (dict, list)):
                    child = node.add(f"{key}")
                    self.populate_json_tree(child, item)
                else:
                    node.add(self.format_json_leaf(f"{key}", item))
        elif isinstance(value, list):
            for idx, item in enumerate(value):
                label = f"[{idx}]"
                if isinstance(item, (dict, list)):
                    child = node.add(label)
                    self.populate_json_tree(child, item)
                else:
                    node.add(self.format_json_leaf(label, item))
        else:
            node.add(self.format_json_leaf("value", value))

    def format_json_leaf(self, key: str, value: Any) -> str:
        type_name = type(value).__name__
        value_text = json.dumps(value, ensure_ascii=False) if not isinstance(value, str) else value
        if len(value_text) > 120:
            value_text = value_text[:117] + "..."
        return f"{key}: {value_text} ({type_name})"

    def xml_to_markdown_preview(self, path: Path, xml_text: str) -> str:
        root_name = "unknown"
        try:
            root = ET.fromstring(xml_text)
            root_name = root.tag
        except Exception:
            pass
        return "\n".join(
            [
                f"# XML Preview: {path.name}",
                "",
                f"- Root tag: `{root_name}`",
                "",
                "```xml",
                xml_text,
                "```",
            ]
        )

    def xml_to_json_obj(self, xml_text: str) -> Any:
        try:
            root = ET.fromstring(xml_text)
        except Exception:
            return {"error": "Failed to parse XML"}
        return {root.tag: self.xml_element_to_obj(root)}

    def xml_element_to_obj(self, element: ET.Element) -> Any:
        obj: dict[str, Any] = {}
        if element.attrib:
            obj["@attributes"] = dict(element.attrib)

        children = list(element)
        if children:
            grouped: dict[str, list[Any]] = {}
            for child in children:
                grouped.setdefault(child.tag, []).append(self.xml_element_to_obj(child))
            for tag, items in grouped.items():
                obj[tag] = items[0] if len(items) == 1 else items

        text = (element.text or "").strip()
        if text:
            if obj:
                obj["#text"] = text
            else:
                return text

        if not obj:
            return ""
        return obj

    def show_image(self, ansi_art: str, meta: str) -> None:
        self.query_one(TabbedContent).active = "image_tab"
        widget = self.query_one("#image_preview", Static)
        combined = f"{meta}\n{ansi_art}" if ansi_art else f"{meta}\n[No image preview available]"
        widget.update(Text.from_ansi(combined))

    def action_render_html(self) -> None:
        if self.current_path is None or self.current_path.suffix.lower() not in HTML_EXTENSIONS:
            self.notify("Render in browser is available only for HTML files.")
            return
        webbrowser.open(self.current_path.resolve().as_uri())
        self.notify(f"Opened in browser: {self.current_path}")

    def action_show_help(self) -> None:
        self.push_screen(HelpScreen())

    def action_search_text(self) -> None:
        self.push_screen(
            TextInputScreen("Search text", "Enter text to search, Esc: cancel", self.last_text_search),
            self.on_search_text_selected,
        )

    def on_search_text_selected(self, query: str | None) -> None:
        if not query:
            return
        self.last_text_search = query
        content = self.get_current_text_content()
        if not content:
            self.notify("No text loaded.")
            return
        self.text_search_matches = self.compute_text_matches(content, query)
        if not self.text_search_matches:
            self.text_search_index = -1
            self.notify("No matches found.")
            return
        self.text_search_index = 0
        self.notify_current_search_match()
        if self.text_search_highlight_all:
            self.render_search_highlight_preview()

    def action_next_match(self) -> None:
        if not self.text_search_matches:
            self.notify("No active search results. Press / to search.")
            return
        self.text_search_index = (self.text_search_index + 1) % len(self.text_search_matches)
        self.notify_current_search_match()
        if self.text_search_highlight_all:
            self.render_search_highlight_preview()

    def action_prev_match(self) -> None:
        if not self.text_search_matches:
            self.notify("No active search results. Press / to search.")
            return
        self.text_search_index = (self.text_search_index - 1) % len(self.text_search_matches)
        self.notify_current_search_match()
        if self.text_search_highlight_all:
            self.render_search_highlight_preview()

    def action_toggle_search_highlight(self) -> None:
        self.text_search_highlight_all = not self.text_search_highlight_all
        if self.text_search_highlight_all:
            self.render_search_highlight_preview()
            self.notify("Search highlight: ON")
        else:
            self.restore_text_rendering()
            self.notify("Search highlight: OFF")

    def action_jump_to_page(self) -> None:
        if self.current_df is None:
            self.notify("Jump to page is available only for table views.")
            return
        total_pages = max(1, (len(self.current_df) + self.table_page_size - 1) // self.table_page_size)
        self.push_screen(
            NumberInputScreen("Jump to page", f"Enter page number (1-{total_pages}), Esc: cancel", str(self.table_page + 1)),
            self.on_jump_page_selected,
        )

    def on_jump_page_selected(self, page: int | None) -> None:
        if page is None or self.current_df is None:
            return
        total_rows = (
            int(
                self.lazy_delimited_context.get("filtered_total_rows")
                if self.lazy_delimited_context is not None and self.lazy_delimited_context.get("filtered_total_rows") is not None
                else self.lazy_delimited_context["total_rows"]
            )
            if self.lazy_delimited_context is not None
            else len(self.current_df)
        )
        total_pages = max(1, (total_rows + self.table_page_size - 1) // self.table_page_size)
        page_index = max(0, min(page - 1, total_pages - 1))
        if page_index != self.table_page:
            self.table_page = page_index
            if self.lazy_delimited_context is not None:
                self.load_lazy_page_async(self.table_page)
            else:
                self.render_current_table_page()

    def action_decrease_page_size(self) -> None:
        self.adjust_page_size(-5)

    def action_increase_page_size(self) -> None:
        self.adjust_page_size(5)

    def adjust_page_size(self, delta: int) -> None:
        if self.current_df is None:
            self.notify("Page size controls are available only for table views.")
            return
        old_size = self.table_page_size
        self.table_page_size = max(5, min(200, self.table_page_size + delta))
        if self.table_page_size == old_size:
            return
        self.table_page = 0
        if self.lazy_delimited_context is not None:
            self.load_lazy_page_async(self.table_page)
        else:
            self.render_current_table_page()
        self.notify(f"Page size: {self.table_page_size}")

    def action_convert_file(self) -> None:
        if self.current_path is None:
            self.notify("Open a file before converting.")
            return
        if self.current_path.suffix.lower() in IMAGE_EXTENSIONS:
            self.notify("Image conversion is not supported here.")
            return
        if self.current_df is None:
            self.notify("Current file is not a convertible table dataset.")
            return
        stem = self.current_path.stem
        suggested = f"{stem}.parquet"
        self.push_screen(ConvertScreen(suggested), self.on_convert_target_selected)

    def action_promote_full_table(self) -> None:
        if self.lazy_delimited_context is None:
            self.notify("Table is already in full mode.")
            return
        ctx = self.lazy_delimited_context
        try:
            df = self.read_delimited_with_fallback(Path(ctx["path"]), sep=ctx["sep"])
            self.lazy_delimited_context = None
            self.current_df_source = df
            self.current_df = df
            if self.table_filter_query.strip():
                self.apply_table_filter()
            else:
                self.apply_sort_if_needed()
            self.table_page = 0
            self.render_current_table_page()
            self.update_status_details()
            self.notify("Promoted to full mode.")
        except Exception as exc:
            self.notify(f"Promote failed: {exc}", severity="error")

    def action_sort_selected_column(self) -> None:
        if self.current_df is None:
            self.notify("No table loaded.")
            return
        if self.lazy_delimited_context is not None:
            self.notify("Sort requires full mode. Press Alt+P (or F6) first.")
            return
        table = self.query_one(DataTable)
        coord = table.cursor_coordinate
        col_index = coord.column
        render_cols = self.get_render_columns(self.current_df)
        if not render_cols:
            self.notify("No visible columns to sort.")
            return
        if col_index < 0 or col_index >= len(render_cols):
            col_index = 0
        col = str(render_cols[col_index])
        asc = True
        if self.current_sort is not None and self.current_sort[0] == col:
            asc = not self.current_sort[1]
        self.current_sort = (col, asc)
        if self.current_df_source is not None:
            self.current_df_source = self.sort_dataframe_by_column(self.current_df_source, col, asc)
        self.apply_table_filter()
        self.table_page = 0
        self.render_current_table_page()
        self.notify(f"Sorted by {col} ({'asc' if asc else 'desc'}).")

    def get_render_columns(self, df: pd.DataFrame) -> list[str]:
        cols = [str(c) for c in df.columns]
        if self.visible_columns:
            selected = [c for c in self.visible_columns if c in cols]
            if selected:
                cols = selected
        return cols[:MAX_RENDER_COLS]

    def action_pick_visible_columns(self) -> None:
        if self.current_df is None:
            self.notify("No table loaded.")
            return
        current = ",".join(self.visible_columns) if self.visible_columns else ""
        hint = "Comma-separated column names. Empty = all columns. Enter to apply."
        self.push_screen(TextInputScreen("Visible columns", hint, current), self.on_visible_columns_selected)

    def on_visible_columns_selected(self, value: str | None) -> None:
        if self.current_df is None:
            return
        if value is None or not value.strip():
            self.visible_columns = None
            self.render_current_table_page()
            self.notify("Showing all columns.")
            return
        requested = [v.strip() for v in value.split(",") if v.strip()]
        if not requested:
            self.visible_columns = None
            self.render_current_table_page()
            self.notify("Showing all columns.")
            return
        available = {str(c): str(c) for c in self.current_df.columns}
        normalized_map = {str(c).lower(): str(c) for c in self.current_df.columns}
        selected: list[str] = []
        missing: list[str] = []
        for col in requested:
            if col in available:
                selected.append(available[col])
            elif col.lower() in normalized_map:
                selected.append(normalized_map[col.lower()])
            else:
                missing.append(col)
        if not selected:
            self.notify("No matching columns found.", severity="warning")
            return
        self.visible_columns = selected
        self.render_current_table_page()
        if missing:
            self.notify(f"Applied columns ({len(selected)}). Missing: {', '.join(missing)}")
        else:
            self.notify(f"Applied columns ({len(selected)}).")

    def on_convert_target_selected(self, value: str | None) -> None:
        if value is None:
            return
        target = Path(value)
        if not target.is_absolute():
            target = Path.cwd() / target
        try:
            self.convert_current_dataframe(target)
            self.refresh_file_tree()
            self.notify(f"Converted: {target}")
            self.query_one("#status", Static).update(f"Converted to: {target}")
        except Exception as exc:
            self.notify(f"Convert failed: {exc}", severity="error")

    def refresh_file_tree(self) -> None:
        tree = self.query_one(ColoredDirectoryTree)
        tree.reload()

    def convert_current_dataframe(self, target: Path) -> None:
        if self.current_df is None:
            raise RuntimeError("No table loaded")
        ext = target.suffix.lower()
        df = self.current_df
        if ext == ".csv":
            df.to_csv(target, index=False)
        elif ext == ".tsv":
            df.to_csv(target, sep="\t", index=False)
        elif ext in {".parquet", ".parq"}:
            df.to_parquet(target, index=False)
        elif ext in {".arrow", ".feather"}:
            df.to_feather(target)
        elif ext == ".orc":
            table = pa.Table.from_pandas(df, preserve_index=False)
            pa_orc.write_table(table, str(target))
        elif ext == ".json":
            df.to_json(target, orient="records", force_ascii=False, indent=2)
        elif ext == ".avro":
            records = df.replace({np.nan: None}).to_dict(orient="records")
            schema = self.infer_avro_schema(df, target.stem or "record")
            with target.open("wb") as fh:
                avro_writer(fh, schema, records)
        else:
            raise ValueError("Unsupported target extension. Use csv, tsv, parquet, parq, arrow, feather, orc, json, avro.")

    def infer_avro_schema(self, df: pd.DataFrame, name: str) -> dict[str, Any]:
        fields: list[dict[str, Any]] = []
        for col in df.columns:
            series = df[col]
            dtype = str(series.dtype)
            if dtype.startswith("int"):
                avro_type: Any = ["null", "long"]
            elif dtype.startswith("float"):
                avro_type = ["null", "double"]
            elif dtype == "bool":
                avro_type = ["null", "boolean"]
            else:
                avro_type = ["null", "string"]
            fields.append({"name": str(col), "type": avro_type})
        return {"type": "record", "name": name, "fields": fields}

    @on(FileLoaded)
    def update_status(self, event: FileLoaded) -> None:
        self.query_one("#status", Static).update(f"Loaded: {event.path}")
        self.update_status_details()

    def update_status_details(self) -> None:
        if self.current_path is None:
            return
        path = self.current_path
        kind = self.current_kind or "unknown"
        suffix = path.suffix.lower() or "[noext]"
        size_text = "unknown size"
        try:
            if path.exists():
                size_text = f"{path.stat().st_size:,} bytes"
        except Exception:
            pass
        extra = ""
        if self.current_df is not None:
            extra = f" | r={len(self.current_df):,} c={len(self.current_df.columns):,}"
            if self.current_df_source is not None and len(self.current_df) != len(self.current_df_source):
                extra += f" (filtered from {len(self.current_df_source):,})"
        mode = "lazy" if self.lazy_delimited_context is not None else "full"
        timing = f" | load={self.last_load_ms}ms"
        cache = f" | cache_hits={self.cache_hits}"
        page_range = ""
        if self.current_df is not None and kind == "table":
            total_rows = (
                int(
                    self.lazy_delimited_context.get("filtered_total_rows")
                    if self.lazy_delimited_context is not None and self.lazy_delimited_context.get("filtered_total_rows") is not None
                    else self.lazy_delimited_context["total_rows"]
                )
                if self.lazy_delimited_context is not None
                else len(self.current_df)
            )
            start_row = self.table_page * self.table_page_size + 1 if total_rows > 0 else 0
            end_row = min(total_rows, (self.table_page + 1) * self.table_page_size)
            page_range = f" | {start_row}-{end_row}/{total_rows}"
        loading = " | page=loading" if self.is_page_loading else ""
        self.query_one("#status", Static).update(
            f"{path} | {suffix} | {kind} | {mode} | {size_text}{extra}{page_range} | {self.last_load_ms}ms{cache}{loading}"
        )

    def action_focus_tree(self) -> None:
        self.query_one(ColoredDirectoryTree).focus()

    async def action_go_parent_dir(self) -> None:
        current_root = self.tree_root
        parent = current_root.parent.resolve()
        if parent == current_root:
            self.notify("Already at filesystem root.")
            return

        self.tree_root = parent
        old_tree = self.query_one("#tree", ColoredDirectoryTree)
        await old_tree.remove()
        main = self.query_one("#main", Horizontal)
        await main.mount(ColoredDirectoryTree(str(self.tree_root), id="tree"), before="#content")
        self.query_one(ColoredDirectoryTree).focus()
        self.query_one("#status", Static).update(f"Explorer root: {self.tree_root}")

    def action_edit_cell(self) -> None:
        if self.lazy_delimited_context is not None:
            self.notify("Cell editing is disabled in lazy CSV/TSV mode.")
            return
        if self.current_kind != "table":
            self.notify("Cell editing is available only for table files.")
            return

        table = self.query_one(DataTable)
        coord = table.cursor_coordinate
        row_index, col_index = coord.row, coord.column
        value = table.get_cell_at(coord)

        def apply_value(new_value: str | None) -> None:
            if new_value is None:
                return
            table.update_cell_at(coord, new_value)
            if self.current_df is not None and self.current_df_source is not None:
                col = self.current_df.columns[col_index]
                actual_row = self.table_page * self.table_page_size + row_index
                if actual_row < len(self.current_df):
                    index_label = self.current_df.index[actual_row]
                    self.current_df.at[index_label, col] = new_value
                    if index_label in self.current_df_source.index:
                        self.current_df_source.at[index_label, col] = new_value

        self.push_screen(CellEditScreen(str(value)), apply_value)

    def action_save(self) -> None:
        if self.current_path is None or self.current_kind is None:
            self.notify("No file open.")
            return

        try:
            if self.current_kind == "table":
                self.save_table(self.current_path)
            elif self.current_kind == "json":
                self.save_json(self.current_path)
            else:
                self.save_text(self.current_path)
            self.notify(f"Saved: {self.current_path}")
            self.query_one("#status", Static).update(f"Saved: {self.current_path}")
        except Exception as exc:
            self.notify(f"Save failed: {exc}", severity="error")

    def save_table(self, path: Path) -> None:
        if self.lazy_delimited_context is not None:
            raise RuntimeError("Save is disabled in lazy CSV/TSV mode. Convert or reload fully first.")
        if self.current_df_source is None:
            raise RuntimeError("No table loaded")
        df_to_save = self.current_df_source
        suffix = path.suffix.lower()
        if suffix == ".csv":
            df_to_save.to_csv(path, index=False)
        elif suffix == ".tsv":
            df_to_save.to_csv(path, index=False, sep="\t")
        elif suffix in {".parquet", ".parq"}:
            df_to_save.to_parquet(path, index=False)
        elif suffix in {".xlsx", ".xls"}:
            df_to_save.to_excel(path, index=False, sheet_name=str(self.current_excel_sheet))
        elif suffix in {".arrow", ".feather"}:
            df_to_save.to_feather(path)
        else:
            raise ValueError("Unsupported table format for save. Use csv, tsv, parquet, parq, xlsx, xls, arrow, or feather.")

    def save_json(self, path: Path) -> None:
        text = self.query_one(TextArea).text
        data: Any = json.loads(text)
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

    def save_text(self, path: Path) -> None:
        text = self.query_one(TextArea).text
        path.write_text(text, encoding="utf-8")

    def prepare_lazy_delimited_context(self, path: Path, sep: str) -> dict[str, Any]:
        encoding = self.detect_text_encoding(path)
        total_rows = self.count_delimited_rows(path, encoding)
        df = pd.read_csv(path, sep=sep, encoding=encoding, nrows=self.table_page_size)
        return {
            "path": path,
            "sep": sep,
            "encoding": encoding,
            "total_rows": max(0, total_rows),
            "df": df,
            "filtered_indices": None,
            "filtered_total_rows": None,
        }

    def load_lazy_page_async(self, page_index: int) -> None:
        if self.lazy_delimited_context is None:
            return
        self.is_page_loading = True
        self.update_status_details()
        self.load_lazy_page_worker(page_index)

    @work(thread=True)
    def load_lazy_page_worker(self, page_index: int) -> None:
        try:
            df = self.read_lazy_delimited_page(page_index)
            self.call_from_thread(self.on_lazy_page_loaded, df)
        except Exception as exc:
            self.call_from_thread(self.notify, f"Lazy page load failed: {exc}", severity="error")
            self.call_from_thread(self.on_lazy_page_loaded, None)

    def on_lazy_page_loaded(self, df: pd.DataFrame | None) -> None:
        if df is not None and self.lazy_delimited_context is not None:
            self.lazy_delimited_context["df"] = df
            if "columns" not in self.lazy_delimited_context:
                self.lazy_delimited_context["columns"] = [str(c) for c in df.columns]
            self.current_df = df
        self.is_page_loading = False
        self.render_current_table_page()
        self.render_schema_profile()
        self.update_status_details()

    def load_lazy_delimited_page(self, page_index: int) -> None:
        df = self.read_lazy_delimited_page(page_index)
        if self.lazy_delimited_context is None:
            return
        self.lazy_delimited_context["df"] = df
        if "columns" not in self.lazy_delimited_context:
            self.lazy_delimited_context["columns"] = [str(c) for c in df.columns]
        self.current_df = df

    def read_lazy_delimited_page(self, page_index: int) -> pd.DataFrame:
        if self.lazy_delimited_context is None:
            return pd.DataFrame()
        ctx = self.lazy_delimited_context
        path: Path = ctx["path"]
        sep = ctx["sep"]
        encoding = ctx["encoding"]
        filtered_indices: list[int] | None = ctx.get("filtered_indices")
        if filtered_indices is None:
            start = max(0, page_index * self.table_page_size)
            skiprows = range(1, start + 1) if start > 0 else None
            df = pd.read_csv(path, sep=sep, encoding=encoding, skiprows=skiprows, nrows=self.table_page_size)
        else:
            start = max(0, page_index * self.table_page_size)
            end = start + self.table_page_size
            wanted = filtered_indices[start:end]
            if not wanted:
                df = pd.DataFrame(columns=ctx.get("columns", []))
            else:
                wanted_set = set(wanted)
                chunks: list[pd.DataFrame] = []
                chunk_size = max(1000, self.table_page_size * 4)
                base_index = 0
                for chunk in pd.read_csv(path, sep=sep, encoding=encoding, chunksize=chunk_size):
                    rel = [idx - base_index for idx in wanted if base_index <= idx < base_index + len(chunk)]
                    if rel:
                        chunks.append(chunk.iloc[rel])
                    base_index += len(chunk)
                    if base_index > wanted[-1]:
                        break
                if chunks:
                    df = pd.concat(chunks, ignore_index=True)
                else:
                    df = pd.DataFrame(columns=ctx.get("columns", []))
        return df

    def apply_lazy_delimited_filter(self, query: str) -> None:
        if self.lazy_delimited_context is None:
            return
        ctx = self.lazy_delimited_context
        path: Path = ctx["path"]
        sep = ctx["sep"]
        encoding = ctx["encoding"]
        chunk_size = max(1000, self.table_page_size * 8)
        matched: list[int] = []
        base_index = 0
        first_chunk = True
        for chunk in pd.read_csv(path, sep=sep, encoding=encoding, chunksize=chunk_size):
            # Normalize per-chunk index so matched positions are always local
            # and can be mapped reliably to absolute file row offsets.
            chunk_local = chunk.reset_index(drop=True)
            filtered_chunk = self.filter_dataframe_with_query(chunk_local, query, notify_errors=first_chunk)
            if not filtered_chunk.empty:
                matched.extend([base_index + int(i) for i in filtered_chunk.index.tolist()])
            base_index += len(chunk_local)
            first_chunk = False
        ctx["filtered_indices"] = matched
        ctx["filtered_total_rows"] = len(matched)
        self.table_page = 0
        self.load_lazy_delimited_page(0)
        self.notify(f"Filter matches: {len(matched)}")

    def detect_text_encoding(self, path: Path) -> str:
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                with path.open("r", encoding=enc) as fh:
                    fh.readline()
                return enc
            except UnicodeDecodeError:
                continue
        return "latin-1"

    def count_delimited_rows(self, path: Path, encoding: str) -> int:
        count = 0
        with path.open("r", encoding=encoding, errors="ignore") as fh:
            for _ in fh:
                count += 1
        return max(0, count - 1)

    def render_schema_profile(self) -> None:
        table = self.query_one("#schema_table", DataTable)
        table.clear(columns=True)
        table.add_column("column")
        table.add_column("dtype")
        table.add_column("nulls")
        table.add_column("sample_unique")
        table.add_column("min")
        table.add_column("max")

        df = self.current_df_source if self.current_df_source is not None else self.current_df
        if df is None or df.empty:
            return

        for col in df.columns[:MAX_RENDER_COLS]:
            series = df[col]
            non_null = series.dropna()
            dtype = str(series.dtype)
            nulls = int(series.isna().sum())

            uniques = pd.unique(non_null.astype(str))
            sample_unique = ", ".join(list(uniques[:3]))
            if len(uniques) > 3:
                sample_unique += ", ..."
            if not sample_unique:
                sample_unique = "-"

            min_text, max_text = self.series_min_max_preview(non_null)
            table.add_row(
                str(col),
                dtype,
                f"{nulls}",
                self.truncate_profile_value(sample_unique, 50),
                self.truncate_profile_value(min_text, 30),
                self.truncate_profile_value(max_text, 30),
            )

    def series_min_max_preview(self, non_null: pd.Series) -> tuple[str, str]:
        if non_null.empty:
            return "-", "-"
        try:
            numeric = pd.to_numeric(non_null, errors="coerce")
            if numeric.notna().sum() > 0:
                return str(numeric.min()), str(numeric.max())
        except Exception:
            pass
        try:
            dt = pd.to_datetime(non_null, errors="coerce")
            if dt.notna().sum() > 0:
                return str(dt.min()), str(dt.max())
        except Exception:
            pass
        text_values = non_null.astype(str)
        return str(text_values.min()), str(text_values.max())

    def truncate_profile_value(self, value: str, limit: int) -> str:
        if len(value) <= limit:
            return value
        return value[: limit - 3] + "..."


def run() -> None:
    FilesViewerApp().run()


if __name__ == "__main__":
    run()
